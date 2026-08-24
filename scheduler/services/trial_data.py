"""Build a deterministic, fictional semester workbook for guided evaluation."""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from scheduler.services.imports import build_import_template

TRIAL_WORKBOOK_FILENAME = "USM-Scheduler-Synthetic-Trial-v1.xlsx"


def _canonicalize_xlsx(content: bytes) -> bytes:
    """Normalize ZIP order and timestamps for a reproducible download."""

    output = BytesIO()
    with ZipFile(BytesIO(content), "r") as source, ZipFile(
        output,
        "w",
        ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = ZipInfo(filename=name, date_time=(2000, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = original.external_attr
            info.internal_attr = original.internal_attr
            info.create_system = 0
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>).*?(</dcterms:modified>)",
                    lambda match: (
                        match.group(1)
                        + b"2000-01-01T00:00:00Z"
                        + match.group(2)
                    ),
                    payload,
                )
            target.writestr(info, payload)
    return output.getvalue()


def _time_atoms() -> list[list[object]]:
    rows: list[list[object]] = []
    for day in range(5):
        cursor = datetime(2000, 1, 1, 8, 0)
        for sequence in range(18):
            next_cursor = cursor + timedelta(minutes=30)
            rows.append(
                [
                    day,
                    sequence,
                    cursor.strftime("%H:%M"),
                    next_cursor.strftime("%H:%M"),
                    sequence in {8, 9},
                    True,
                ]
            )
            cursor = next_cursor
    return rows


def build_trial_workbook_bytes(*, campus: str = "Kabacan") -> bytes:
    """Return a realistic but wholly synthetic one-campus scheduling dataset.

    The workbook deliberately covers the rules most useful in a thesis trial:
    contextual subject classifications, explicit room grants, laboratories,
    a shared offering, team teaching, restricted availability, preferences,
    repeated meetings on distinct days, and one valid locked placement.
    """

    workbook = load_workbook(BytesIO(build_import_template()))
    fixed_timestamp = datetime(2000, 1, 1)
    workbook.properties.created = fixed_timestamp
    workbook.properties.modified = fixed_timestamp
    workbook.properties.title = "USM Scheduler Synthetic Trial Dataset"
    workbook.properties.subject = "Fictional, de-identified data for thesis-prototype testing"
    workbook.properties.description = (
        "SYNTHETIC TEST DATA ONLY. Names, codes, rooms, availability, and assignments "
        "do not represent official University of Southern Mindanao records."
    )
    workbook.properties.keywords = "synthetic,test,university timetabling,USM thesis prototype"
    workbook.properties.creator = "USM Scheduler thesis prototype"

    rows: dict[str, list[list[object]]] = {
        "Colleges": [
            ["SYN-CEC", "Synthetic College of Computing", True],
            ["SYN-CSM", "Synthetic College of Science and Mathematics", True],
            ["SYN-CASS", "Synthetic College of Arts and Social Sciences", True],
        ],
        "Departments": [
            ["SYN-DCS", "Synthetic Department of Computer Science", "SYN-CEC", True],
            ["SYN-DIT", "Synthetic Department of Information Technology", "SYN-CEC", True],
            ["SYN-DMATH", "Synthetic Department of Mathematics", "SYN-CSM", True],
            ["SYN-DCOMM", "Synthetic General Education Unit", "SYN-CASS", True],
        ],
        "Programs": [
            ["SYN-BSCS", "Synthetic BS Computer Science", "SYN-DCS", "2026-TEST", True],
            ["SYN-BSIT", "Synthetic BS Information Technology", "SYN-DIT", "2026-TEST", True],
        ],
        "Subjects": [
            ["TCS101", "Foundations of Computing", "Synthetic major lecture", True],
            ["TCS102", "Programming Fundamentals", "Synthetic lecture and laboratory", True],
            ["TCS201", "Data Structures", "Synthetic twice-weekly major", True],
            ["TCS401", "Computing Project", "Synthetic team-taught major", True],
            ["TIT101", "Information Technology Fundamentals", "Synthetic IT major", True],
            ["TMTH101", "Mathematics for Computing", "Synthetic service subject", True],
            ["TGE101", "Communication in Context", "Synthetic shared GE subject", True],
        ],
        "ProgramSubjects": [
            ["SYN-BSCS", "TCS101", "2026-TEST", "MAJOR", "SYN-CEC", "SYN-DCS", True],
            ["SYN-BSCS", "TCS102", "2026-TEST", "MAJOR", "SYN-CEC", "SYN-DCS", True],
            ["SYN-BSCS", "TCS201", "2026-TEST", "MAJOR", "SYN-CEC", "SYN-DCS", True],
            ["SYN-BSCS", "TCS401", "2026-TEST", "MAJOR", "SYN-CEC", "SYN-DCS", True],
            ["SYN-BSIT", "TIT101", "2026-TEST", "MAJOR", "SYN-CEC", "SYN-DIT", True],
            ["SYN-BSCS", "TMTH101", "2026-TEST", "MINOR", "SYN-CSM", "SYN-DMATH", True],
            ["SYN-BSCS", "TGE101", "2026-TEST", "GE", "SYN-CASS", "SYN-DCOMM", True],
            ["SYN-BSIT", "TGE101", "2026-TEST", "GE", "SYN-CASS", "SYN-DCOMM", True],
        ],
        "Sections": [
            ["SYN-BSCS-1A", "SYN-BSCS", 1, "INCOMING", True],
            ["SYN-BSCS-1B", "SYN-BSCS", 1, "INCOMING", True],
            ["SYN-BSCS-2A", "SYN-BSCS", 2, "CONTINUING", True],
            ["SYN-BSCS-4A", "SYN-BSCS", 4, "GRADUATING", True],
            ["SYN-BSIT-1A", "SYN-BSIT", 1, "INCOMING", True],
        ],
        "Instructors": [
            ["SYN-FAC-001", "Synthetic Faculty 01", "SYN-DCS", True, True],
            ["SYN-FAC-002", "Synthetic Faculty 02", "SYN-DCS", True, True],
            ["SYN-FAC-003", "Synthetic Faculty 03", "SYN-DCS", True, True],
            ["SYN-FAC-004", "Synthetic Faculty 04", "SYN-DIT", True, False],
            ["SYN-FAC-005", "Synthetic Faculty 05", "SYN-DCOMM", True, True],
            ["SYN-FAC-006", "Synthetic Faculty 06", "SYN-DMATH", True, True],
        ],
        "Rooms": [
            ["TEST-CEC-101", "Synthetic Computing Classroom 101", campus, "CLASSROOM", "", "SYN-DCS", True, True],
            ["TEST-CEC-102", "Synthetic Computing Classroom 102", campus, "CLASSROOM", "", "SYN-DIT", True, True],
            ["TEST-CEC-LAB1", "Synthetic Computer Laboratory 1", campus, "LABORATORY", "", "SYN-DCS", True, True],
            ["TEST-CEC-LAB2", "Synthetic Computer Laboratory 2", campus, "LABORATORY", "", "SYN-DCS", True, False],
            ["TEST-CSM-201", "Synthetic Mathematics Room 201", campus, "CLASSROOM", "", "SYN-DMATH", True, True],
            ["TEST-CASS-101", "Synthetic GE Classroom 101", campus, "CLASSROOM", "", "SYN-DCOMM", True, True],
            ["TEST-USM-AVR", "Synthetic Shared Presentation Room", campus, "SPECIAL", "SYN-CASS", "", True, True],
        ],
        "Capabilities": [
            ["COMPUTER_LAB", "Computer laboratory", "Synthetic computing-lab capability"],
            ["PRESENTATION", "Presentation equipment", "Synthetic audiovisual capability"],
        ],
        "RoomCapabilities": [
            ["TEST-CEC-LAB1", "COMPUTER_LAB"],
            ["TEST-CEC-LAB2", "COMPUTER_LAB"],
            ["TEST-USM-AVR", "PRESENTATION"],
        ],
        "RoomAuthorizations": [
            ["TEST-CEC-101", "MAJOR", "", "SYN-DCS", "Explicit DCS major grant"],
            ["TEST-CEC-101", "MAJOR", "", "SYN-DIT", "Shared with synthetic DIT"],
            ["TEST-CEC-102", "MAJOR", "", "SYN-DCS", "Shared with synthetic DCS"],
            ["TEST-CEC-102", "MAJOR", "", "SYN-DIT", "Explicit DIT major grant"],
            ["TEST-CEC-LAB1", "MAJOR", "", "SYN-DCS", "DCS laboratory grant"],
            ["TEST-CEC-LAB2", "MAJOR", "", "SYN-DCS", "DCS laboratory grant"],
            ["TEST-CSM-201", "MINOR", "", "SYN-DMATH", "Mathematics offering-unit grant"],
            ["TEST-USM-AVR", "MINOR", "", "SYN-DMATH", "Borrowed service-room grant"],
            ["TEST-CASS-101", "GE", "", "SYN-DCOMM", "GE offering-unit grant"],
            ["TEST-USM-AVR", "GE", "", "SYN-DCOMM", "Shared GE room grant"],
        ],
        "TimeSlots": _time_atoms(),
        "CourseOfferings": [
            ["SYN-CS101-1A", "TCS101", "SYN-DCS", True],
            ["SYN-CS101-1B", "TCS101", "SYN-DCS", True],
            ["SYN-CS102-1A", "TCS102", "SYN-DCS", True],
            ["SYN-CS102-1B", "TCS102", "SYN-DCS", True],
            ["SYN-CS201-2A", "TCS201", "SYN-DCS", True],
            ["SYN-CS401-4A", "TCS401", "SYN-DCS", True],
            ["SYN-IT101-1A", "TIT101", "SYN-DIT", True],
            ["SYN-MTH101-1A", "TMTH101", "SYN-DMATH", True],
            ["SYN-MTH101-1B", "TMTH101", "SYN-DMATH", True],
            ["SYN-GE101-SHARED", "TGE101", "SYN-DCOMM", True],
            ["SYN-GE101-2A", "TGE101", "SYN-DCOMM", True],
        ],
        "OfferingSections": [
            ["SYN-CS101-1A", "SYN-BSCS-1A", "SYN-BSCS", "TCS101", "2026-TEST"],
            ["SYN-CS101-1B", "SYN-BSCS-1B", "SYN-BSCS", "TCS101", "2026-TEST"],
            ["SYN-CS102-1A", "SYN-BSCS-1A", "SYN-BSCS", "TCS102", "2026-TEST"],
            ["SYN-CS102-1B", "SYN-BSCS-1B", "SYN-BSCS", "TCS102", "2026-TEST"],
            ["SYN-CS201-2A", "SYN-BSCS-2A", "SYN-BSCS", "TCS201", "2026-TEST"],
            ["SYN-CS401-4A", "SYN-BSCS-4A", "SYN-BSCS", "TCS401", "2026-TEST"],
            ["SYN-IT101-1A", "SYN-BSIT-1A", "SYN-BSIT", "TIT101", "2026-TEST"],
            ["SYN-MTH101-1A", "SYN-BSCS-1A", "SYN-BSCS", "TMTH101", "2026-TEST"],
            ["SYN-MTH101-1B", "SYN-BSCS-1B", "SYN-BSCS", "TMTH101", "2026-TEST"],
            ["SYN-GE101-SHARED", "SYN-BSCS-1A", "SYN-BSCS", "TGE101", "2026-TEST"],
            ["SYN-GE101-SHARED", "SYN-BSIT-1A", "SYN-BSIT", "TGE101", "2026-TEST"],
            ["SYN-GE101-2A", "SYN-BSCS-2A", "SYN-BSCS", "TGE101", "2026-TEST"],
        ],
        "OfferingInstructors": [
            ["SYN-CS101-1A", "SYN-FAC-001"],
            ["SYN-CS101-1B", "SYN-FAC-001"],
            ["SYN-CS102-1A", "SYN-FAC-002"],
            ["SYN-CS102-1B", "SYN-FAC-002"],
            ["SYN-CS201-2A", "SYN-FAC-003"],
            ["SYN-CS401-4A", "SYN-FAC-003"],
            ["SYN-CS401-4A", "SYN-FAC-004"],
            ["SYN-IT101-1A", "SYN-FAC-004"],
            ["SYN-MTH101-1A", "SYN-FAC-006"],
            ["SYN-MTH101-1B", "SYN-FAC-006"],
            ["SYN-GE101-SHARED", "SYN-FAC-005"],
            ["SYN-GE101-2A", "SYN-FAC-005"],
        ],
        "MeetingRequirements": [
            ["SYN-CS101-1A-LEC", "SYN-CS101-1A", "LECTURE", 1, 3, "", True],
            ["SYN-CS101-1B-LEC", "SYN-CS101-1B", "LECTURE", 1, 3, "", True],
            ["SYN-CS102-1A-LEC", "SYN-CS102-1A", "LECTURE", 1, 2, "", True],
            ["SYN-CS102-1A-LAB", "SYN-CS102-1A", "LAB", 1, 4, "", True],
            ["SYN-CS102-1B-LEC", "SYN-CS102-1B", "LECTURE", 1, 2, "", True],
            ["SYN-CS102-1B-LAB", "SYN-CS102-1B", "LAB", 1, 4, "", True],
            ["SYN-CS201-2A-LEC1", "SYN-CS201-2A", "LECTURE", 1, 2, "TWICE-WEEKLY", True],
            ["SYN-CS201-2A-LEC2", "SYN-CS201-2A", "LECTURE", 2, 2, "TWICE-WEEKLY", True],
            ["SYN-CS401-4A-PROJ", "SYN-CS401-4A", "LECTURE", 1, 4, "", True],
            ["SYN-IT101-1A-LEC", "SYN-IT101-1A", "LECTURE", 1, 3, "", True],
            ["SYN-MTH101-1A-LEC", "SYN-MTH101-1A", "LECTURE", 1, 3, "", True],
            ["SYN-MTH101-1B-LEC", "SYN-MTH101-1B", "LECTURE", 1, 3, "", True],
            ["SYN-GE101-SHARED-LEC", "SYN-GE101-SHARED", "LECTURE", 1, 3, "", True],
            ["SYN-GE101-2A-LEC", "SYN-GE101-2A", "LECTURE", 1, 3, "", True],
        ],
        "MeetingCapabilities": [
            ["SYN-CS102-1A-LAB", "COMPUTER_LAB"],
            ["SYN-CS102-1B-LAB", "COMPUTER_LAB"],
        ],
        "InstructorAvailability": [],
        "RoomAvailability": [],
        "Students": [],
        "StudentSections": [],
        "InstructorPreferences": [
            ["SYN-FAC-001", 0, 0, "PREFERRED", 2],
            ["SYN-FAC-001", 4, 14, "AVOID", 1],
            ["SYN-FAC-002", 1, 0, "PREFERRED", 2],
            ["SYN-FAC-003", 0, 10, "PREFERRED", 1],
            ["SYN-FAC-005", 4, 14, "AVOID", 2],
            ["SYN-FAC-006", 2, 0, "PREFERRED", 1],
        ],
        "LaboratoryProfiles": [
            ["TEST-CEC-LAB1", "Synthetic computer laboratory", "Trial data only"],
            ["TEST-CEC-LAB2", "Synthetic computer laboratory", "Restricted to Tuesday and Thursday"],
        ],
        "Locks": [
            [
                "SYN-CS401-4A-PROJ",
                "TEST-CEC-101",
                0,
                10,
                "Synthetic approved project block for lock testing",
            ]
        ],
    }

    # Synthetic Faculty 04 is available on Monday, Wednesday, and Friday only.
    for day in (0, 2, 4):
        for sequence in (*range(0, 8), *range(10, 18)):
            rows["InstructorAvailability"].append(["SYN-FAC-004", day, sequence, True])

    # The second laboratory models a planned Tuesday/Thursday availability profile.
    for day in (1, 3):
        for sequence in (*range(0, 8), *range(10, 18)):
            rows["RoomAvailability"].append(["TEST-CEC-LAB2", day, sequence, True])

    for sheet_name, sheet_rows in rows.items():
        sheet = workbook[sheet_name]
        sheet.sheet_properties.tabColor = "0E5A3B"
        for row in sheet_rows:
            sheet.append(row)
        if sheet.max_row > 1:
            sheet.auto_filter.ref = sheet.dimensions

    # Repeat a visible synthetic-data banner above each populated data table via
    # comments would interfere with the strict schema. Workbook metadata and the
    # distinctive SYN-/TEST- identifiers instead keep the status machine-readable.
    schema_sheet = workbook["_Schema"]
    schema_sheet["A3"] = "dataset_notice"
    schema_sheet["B3"] = "SYNTHETIC TEST DATA ONLY - NOT OFFICIAL USM RECORDS"
    schema_sheet["A3"].font = Font(bold=True, color="9C2C2C")
    schema_sheet["B3"].font = Font(bold=True, color="9C2C2C")
    schema_sheet["A3"].fill = PatternFill(fill_type="solid", fgColor="FFF1F0")
    schema_sheet["B3"].fill = PatternFill(fill_type="solid", fgColor="FFF1F0")

    output = BytesIO()
    workbook.save(output)
    return _canonicalize_xlsx(output.getvalue())


__all__ = ["TRIAL_WORKBOOK_FILENAME", "build_trial_workbook_bytes"]
