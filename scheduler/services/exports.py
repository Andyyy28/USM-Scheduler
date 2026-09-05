"""Schedule exports without student identity, plus reproducibility manifests."""

from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from typing import Any

from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scheduler import models
from scheduler.services.assignment_display import prepare_assignments

SCHEDULE_COLUMNS = (
    "meeting_id",
    "offering_key",
    "subject_code",
    "subject_title",
    "component",
    "sections",
    "instructors",
    "day",
    "starts_at",
    "ends_at",
    "room_code",
    "offering_unit",
    "locked",
)


def _schedule_queryset(schedule: models.ScheduleVersion):
    return schedule.assignments.select_related(
        "meeting_requirement__offering__subject",
        "meeting_requirement__offering__offering_department__college",
        "room",
        "start_time_slot",
    ).prefetch_related(
        "meeting_requirement__offering__section_links__section",
        "meeting_requirement__offering__instructor_links__instructor",
        "room_allocations__time_slot",
    )


def schedule_export_rows(schedule: models.ScheduleVersion) -> list[dict[str, Any]]:
    """Return stable section-level rows without student-identifying data."""

    assignments = prepare_assignments(schedule, list(_schedule_queryset(schedule)))
    meeting_ids = [row.meeting_requirement_id for row in assignments]
    locked_meetings = set(
        models.LockedAssignment.objects.filter(
            meeting_requirement_id__in=meeting_ids,
            is_active=True,
        ).values_list("meeting_requirement_id", flat=True)
    )
    rows: list[dict[str, Any]] = []
    for assignment in assignments:
        meeting = assignment.meeting_requirement
        offering = meeting.offering
        section_codes = sorted(
            link.section.code for link in offering.section_links.all() if link.section.is_active
        )
        instructors = sorted(
            f"{link.instructor.employee_code} - {link.instructor.display_name}"
            for link in offering.instructor_links.all()
            if link.instructor.is_active
        )
        occupied = sorted(
            (allocation.time_slot for allocation in assignment.room_allocations.all()),
            key=lambda slot: (slot.day, slot.sequence),
        )
        end_time = getattr(assignment, "resolved_end_time", None)
        if not schedule.snapshot_id:
            end_time = occupied[-1].ends_at if occupied else assignment.start_time_slot.ends_at
        rows.append(
            {
                "meeting_id": str(meeting.stable_key),
                "offering_key": offering.external_key,
                "subject_code": offering.subject.code,
                "subject_title": offering.subject.title,
                "component": meeting.component,
                "sections": "; ".join(section_codes),
                "instructors": "; ".join(instructors),
                "day": assignment.start_time_slot.get_day_display(),
                "starts_at": assignment.start_time_slot.starts_at.strftime("%H:%M"),
                "ends_at": end_time.strftime("%H:%M") if end_time else "Unresolved placement",
                "room_code": assignment.room.code,
                "offering_unit": (
                    f"{offering.offering_department.college.code} / "
                    f"{offering.offering_department.code}"
                ),
                "locked": "YES" if getattr(assignment, "resolved_locked", meeting.pk in locked_meetings) else "NO",
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row["day"],
            row["starts_at"],
            row["subject_code"],
            row["sections"],
        ),
    )


def schedule_export_filename(schedule: models.ScheduleVersion, extension: str) -> str:
    stem = slugify(
        f"{schedule.term.academic_year}-{schedule.term.semester}-"
        f"{schedule.term.campus}-schedule-v{schedule.version_number}"
    )
    return f"{stem}.{extension}"


def schedule_csv_bytes(schedule: models.ScheduleVersion) -> bytes:
    stream = StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=SCHEDULE_COLUMNS)
    writer.writeheader()
    writer.writerows(schedule_export_rows(schedule))
    return stream.getvalue().encode("utf-8-sig")


def _metadata_rows(schedule: models.ScheduleVersion) -> list[tuple[str, Any]]:
    validation = getattr(schedule, "validation_result", None)
    return [
        ("Schedule ID", schedule.pk),
        ("Name", schedule.name),
        ("Academic year", schedule.term.academic_year),
        ("Semester", schedule.term.get_semester_display()),
        ("Campus", schedule.term.campus),
        ("Version", schedule.version_number),
        ("Status", schedule.get_status_display()),
        ("Dataset revision", schedule.revision.revision_number),
        ("Dataset hash", schedule.revision.content_hash),
        ("Problem hash", schedule.snapshot.snapshot_hash if schedule.snapshot_id else ""),
        ("Objective profile hash", schedule.snapshot.objective_profile.profile_hash if schedule.snapshot_id else ""),
        ("Objective penalty", schedule.objective_value),
        ("Independently feasible", validation.is_feasible if validation else "Not validated"),
        ("Hard violations", validation.hard_violation_count if validation else ""),
        ("Finalized at", schedule.finalized_at.isoformat() if schedule.finalized_at else ""),
    ]


def schedule_xlsx_bytes(schedule: models.ScheduleVersion) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    header_fill = PatternFill("solid", fgColor="14532D")
    for column, name in enumerate(SCHEDULE_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=name.replace("_", " ").title())
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for row_number, row in enumerate(schedule_export_rows(schedule), start=2):
        for column, name in enumerate(SCHEDULE_COLUMNS, start=1):
            sheet.cell(row=row_number, column=column, value=row[name])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, name in enumerate(SCHEDULE_COLUMNS, start=1):
        widths = [len(str(name))] + [
            len(str(sheet.cell(row=row, column=column).value or ""))
            for row in range(2, min(sheet.max_row, 250) + 1)
        ]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(widths) + 2, 48)

    metadata = workbook.create_sheet("Manifest")
    metadata.append(["Field", "Value"])
    metadata["A1"].font = metadata["B1"].font = Font(bold=True, color="FFFFFF")
    metadata["A1"].fill = metadata["B1"].fill = header_fill
    for key, value in _metadata_rows(schedule):
        metadata.append([key, value])
    metadata.column_dimensions["A"].width = 28
    metadata.column_dimensions["B"].width = 72

    validation = getattr(schedule, "validation_result", None)
    if validation:
        validation_sheet = workbook.create_sheet("Validation")
        validation_sheet.append(["Independent validation report (JSON)"])
        validation_sheet["A1"].font = Font(bold=True)
        validation_sheet.append(
            [json.dumps(validation.violations, ensure_ascii=False, sort_keys=True, indent=2)]
        )
        validation_sheet.column_dimensions["A"].width = 100

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def snapshot_manifest_bytes(snapshot: models.ProblemSnapshot) -> bytes:
    manifest = {
        "schema_version": snapshot.schema_version,
        "snapshot_id": snapshot.pk,
        "snapshot_hash": snapshot.snapshot_hash,
        "dataset_revision_id": snapshot.revision_id,
        "dataset_revision_hash": snapshot.revision.content_hash,
        "objective_profile_id": snapshot.objective_profile_id,
        "objective_profile_hash": snapshot.objective_profile.profile_hash,
        "event_count": snapshot.event_count,
        "candidate_count": snapshot.candidate_count,
        "preprocessing_seconds": snapshot.preprocessing_seconds,
        "created_at": snapshot.created_at.isoformat(),
    }
    return json.dumps(manifest, indent=2, sort_keys=True).encode("utf-8")
