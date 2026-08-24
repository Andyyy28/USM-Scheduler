"""Validated, versioned XLSX import pipeline for semester scheduling data.

Raw workbook bytes are parsed in memory. Only normalized JSON-compatible rows and
cell-level validation errors are persisted; uploaded workbooks are never written
to the filesystem.
"""

from __future__ import annotations

import hashlib
import uuid
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime, time
from io import BytesIO
from typing import Any, Literal

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Max
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from scheduler import models

SCHEMA_VERSION = "1.0"
MAX_WORKBOOK_BYTES = 20 * 1024 * 1024
MAX_ROWS_PER_SHEET = 100_000
SCHEMA_SHEET = "_Schema"

CellKind = Literal["string", "integer", "boolean", "time", "day"]


class ImportPipelineError(ValueError):
    """Base exception for import preview and commit failures."""


class ImportCommitError(ImportPipelineError):
    """Raised when a preview cannot be committed safely."""


@dataclass(frozen=True, slots=True)
class ColumnSpec:
    name: str
    kind: CellKind = "string"
    required: bool = True
    choices: tuple[str, ...] = ()
    description: str = ""


@dataclass(frozen=True, slots=True)
class SheetSpec:
    columns: tuple[ColumnSpec, ...]
    key: tuple[str, ...]
    required: bool = True


def _column(
    name: str,
    kind: CellKind = "string",
    *,
    required: bool = True,
    choices: Iterable[str] = (),
    description: str = "",
) -> ColumnSpec:
    return ColumnSpec(name, kind, required, tuple(choices), description)


BOOL = _column("", "boolean", required=False)
CLASSIFICATION_CHOICES = tuple(models.SubjectClassification.values)


SHEET_SCHEMAS: dict[str, SheetSpec] = {
    "Colleges": SheetSpec(
        (
            _column("code", description="Unique college code."),
            _column("name"),
            _column("is_active", "boolean", required=False),
        ),
        ("code",),
    ),
    "Departments": SheetSpec(
        (
            _column("code"),
            _column("name"),
            _column("college_code"),
            _column("is_active", "boolean", required=False),
        ),
        ("code",),
    ),
    "Programs": SheetSpec(
        (
            _column("code"),
            _column("name"),
            _column("department_code"),
            _column("curriculum_label", required=False),
            _column("is_active", "boolean", required=False),
        ),
        ("code",),
    ),
    "Subjects": SheetSpec(
        (
            _column("code"),
            _column("title"),
            _column("description", required=False),
            _column("is_active", "boolean", required=False),
        ),
        ("code",),
    ),
    "ProgramSubjects": SheetSpec(
        (
            _column("program_code"),
            _column("subject_code"),
            _column("curriculum_version"),
            _column("classification", choices=CLASSIFICATION_CHOICES),
            _column("authoritative_college_code"),
            _column("authoritative_department_code", required=False),
            _column("is_active", "boolean", required=False),
        ),
        ("program_code", "subject_code", "curriculum_version"),
    ),
    "Sections": SheetSpec(
        (
            _column("code"),
            _column("program_code"),
            _column("year_level", "integer"),
            _column("cohort_status", choices=tuple(models.CohortStatus.values)),
            _column("is_active", "boolean", required=False),
        ),
        ("code",),
    ),
    "Instructors": SheetSpec(
        (
            _column("employee_code"),
            _column("display_name"),
            _column("department_code"),
            _column("is_active", "boolean", required=False),
            _column("assume_fully_available", "boolean", required=False),
        ),
        ("employee_code",),
    ),
    "Rooms": SheetSpec(
        (
            _column("code"),
            _column("name", required=False),
            _column("campus"),
            _column("kind", choices=tuple(models.RoomKind.values)),
            _column("owning_college_code", required=False),
            _column("owning_department_code", required=False),
            _column("is_active", "boolean", required=False),
            _column("assume_fully_available", "boolean", required=False),
        ),
        ("code",),
    ),
    "Capabilities": SheetSpec(
        (
            _column("code"),
            _column("name"),
            _column("description", required=False),
        ),
        ("code",),
    ),
    "RoomCapabilities": SheetSpec(
        (_column("room_code"), _column("capability_code")),
        ("room_code", "capability_code"),
    ),
    "RoomAuthorizations": SheetSpec(
        (
            _column("room_code"),
            _column("classification", choices=CLASSIFICATION_CHOICES),
            _column("college_code", required=False),
            _column("department_code", required=False),
            _column("notes", required=False),
        ),
        ("room_code", "classification", "college_code", "department_code"),
    ),
    "TimeSlots": SheetSpec(
        (
            _column("day", "day"),
            _column("sequence", "integer"),
            _column("starts_at", "time"),
            _column("ends_at", "time"),
            _column("is_break", "boolean", required=False),
            _column("is_active", "boolean", required=False),
        ),
        ("day", "sequence"),
    ),
    "CourseOfferings": SheetSpec(
        (
            _column("external_key"),
            _column("subject_code"),
            _column("offering_department_code"),
            _column("is_active", "boolean", required=False),
        ),
        ("external_key",),
    ),
    "OfferingSections": SheetSpec(
        (
            _column("offering_key"),
            _column("section_code"),
            _column("program_code"),
            _column("subject_code"),
            _column("curriculum_version"),
        ),
        ("offering_key", "section_code"),
    ),
    "OfferingInstructors": SheetSpec(
        (_column("offering_key"), _column("instructor_code")),
        ("offering_key", "instructor_code"),
    ),
    "MeetingRequirements": SheetSpec(
        (
            _column("meeting_key"),
            _column("offering_key"),
            _column("component", choices=tuple(models.MeetingComponent.values)),
            _column("occurrence_number", "integer"),
            _column("duration_atoms", "integer"),
            _column("distinct_day_group", required=False),
            _column("is_active", "boolean", required=False),
        ),
        ("meeting_key",),
    ),
    "MeetingCapabilities": SheetSpec(
        (_column("meeting_key"), _column("capability_code")),
        ("meeting_key", "capability_code"),
    ),
    "InstructorAvailability": SheetSpec(
        (
            _column("instructor_code"),
            _column("day", "day"),
            _column("sequence", "integer"),
            _column("is_available", "boolean"),
        ),
        ("instructor_code", "day", "sequence"),
    ),
    "RoomAvailability": SheetSpec(
        (
            _column("room_code"),
            _column("day", "day"),
            _column("sequence", "integer"),
            _column("is_available", "boolean"),
        ),
        ("room_code", "day", "sequence"),
    ),
    "Students": SheetSpec(
        (
            _column("pseudonymous_code"),
            _column("status", choices=tuple(models.StudentStatus.values)),
        ),
        ("pseudonymous_code",),
        required=False,
    ),
    "StudentSections": SheetSpec(
        (_column("pseudonymous_code"), _column("section_code")),
        ("pseudonymous_code", "section_code"),
        required=False,
    ),
    "InstructorPreferences": SheetSpec(
        (
            _column("instructor_code"),
            _column("day", "day"),
            _column("sequence", "integer"),
            _column("level", choices=tuple(models.PreferenceLevel.values)),
            _column("weight", "integer"),
        ),
        ("instructor_code", "day", "sequence"),
        required=False,
    ),
    "LaboratoryProfiles": SheetSpec(
        (
            _column("room_code"),
            _column("laboratory_type"),
            _column("notes", required=False),
        ),
        ("room_code",),
        required=False,
    ),
    "Locks": SheetSpec(
        (
            _column("meeting_key"),
            _column("room_code"),
            _column("day", "day"),
            _column("sequence", "integer"),
            _column("reason"),
        ),
        ("meeting_key",),
        required=False,
    ),
}

CORE_SHEETS = tuple(name for name, schema in SHEET_SCHEMAS.items() if schema.required)
OPTIONAL_SHEETS = tuple(name for name, schema in SHEET_SCHEMAS.items() if not schema.required)

DAY_NAMES = {
    "MONDAY": 0,
    "MON": 0,
    "TUESDAY": 1,
    "TUE": 1,
    "TUES": 1,
    "WEDNESDAY": 2,
    "WED": 2,
    "THURSDAY": 3,
    "THU": 3,
    "THUR": 3,
    "FRIDAY": 4,
    "FRI": 4,
    "SATURDAY": 5,
    "SAT": 5,
    "SUNDAY": 6,
    "SUN": 6,
}


@dataclass(frozen=True, slots=True)
class ParsedRow:
    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class Issue:
    sheet_name: str
    row_number: int | None
    column_name: str
    code: str
    message: str


def _issue(
    issues: list[Issue],
    sheet: str,
    row: int | None,
    column: str,
    code: str,
    message: str,
) -> None:
    issues.append(Issue(sheet, row, column, code, message))


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    normalized = str(value).strip().upper()
    if normalized in {"TRUE", "YES", "Y", "1"}:
        return True
    if normalized in {"FALSE", "NO", "N", "0"}:
        return False
    raise ValueError("Use TRUE/FALSE, YES/NO, or 1/0.")


def _parse_integer(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError("A Boolean value is not an integer here.")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if text.lstrip("+-").isdigit():
        return int(text)
    raise ValueError("Enter a whole number.")


def _parse_time(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0).strftime("%H:%M")
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M")
        except ValueError:
            continue
    raise ValueError("Enter a time such as 08:00 or 1:30 PM.")


def _parse_day(value: Any) -> int:
    try:
        day = _parse_integer(value)
    except ValueError:
        normalized = str(value).strip().upper()
        if normalized not in DAY_NAMES:
            raise ValueError("Use Monday-Sunday or an integer from 0 to 6.") from None
        day = DAY_NAMES[normalized]
    if day not in range(7):
        raise ValueError("Day must be between 0 (Monday) and 6 (Sunday).")
    return day


def _normalize_cell(value: Any, spec: ColumnSpec) -> Any:
    if _is_blank(value):
        if spec.required:
            raise ValueError("This value is required.")
        return None
    if spec.kind == "boolean":
        normalized: Any = _parse_boolean(value)
    elif spec.kind == "integer":
        normalized = _parse_integer(value)
    elif spec.kind == "time":
        normalized = _parse_time(value)
    elif spec.kind == "day":
        normalized = _parse_day(value)
    else:
        normalized = str(value).strip()
    if spec.choices:
        choice = str(normalized).upper()
        if choice not in spec.choices:
            raise ValueError(f"Choose one of: {', '.join(spec.choices)}.")
        normalized = choice
    return normalized


def _read_schema_version(workbook: Any, issues: list[Issue]) -> str | None:
    if SCHEMA_SHEET not in workbook.sheetnames:
        _issue(
            issues,
            SCHEMA_SHEET,
            None,
            "schema_version",
            "MISSING_SCHEMA_VERSION",
            f"Workbook is missing the {SCHEMA_SHEET} version sheet.",
        )
        return None
    sheet = workbook[SCHEMA_SHEET]
    key = sheet["A1"].value
    value = sheet["B1"].value
    if key != "schema_version" or str(value).strip() != SCHEMA_VERSION:
        _issue(
            issues,
            SCHEMA_SHEET,
            1,
            "schema_version",
            "UNSUPPORTED_SCHEMA_VERSION",
            f"Expected schema version {SCHEMA_VERSION}; received {value!r}.",
        )
        return str(value).strip() if value is not None else None
    return SCHEMA_VERSION


def _parse_sheet(workbook: Any, name: str, schema: SheetSpec, issues: list[Issue]) -> list[ParsedRow]:
    if name not in workbook.sheetnames:
        if schema.required:
            _issue(issues, name, None, "", "MISSING_SHEET", f"Required sheet {name!r} is missing.")
        return []
    sheet = workbook[name]
    if sheet.max_row > MAX_ROWS_PER_SHEET + 1:
        _issue(
            issues,
            name,
            None,
            "",
            "TOO_MANY_ROWS",
            f"Sheet exceeds the {MAX_ROWS_PER_SHEET:,}-row safety limit.",
        )
        return []

    header_cells = list(sheet[1])
    headers: list[str] = []
    for cell in header_cells:
        if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
            _issue(issues, name, 1, cell.coordinate, "FORMULA_NOT_ALLOWED", "Formulas are not allowed.")
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    while headers and not headers[-1]:
        headers.pop()

    duplicates = sorted({header for header in headers if header and headers.count(header) > 1})
    for header in duplicates:
        _issue(issues, name, 1, header, "DUPLICATE_HEADER", f"Header {header!r} appears more than once.")
    expected = [column.name for column in schema.columns]
    missing = [header for header in expected if header not in headers]
    unexpected = [header for header in headers if header and header not in expected]
    for header in missing:
        _issue(issues, name, 1, header, "MISSING_HEADER", f"Required header {header!r} is missing.")
    for header in unexpected:
        _issue(issues, name, 1, header, "UNEXPECTED_HEADER", f"Unexpected header {header!r}.")
    if duplicates or missing or unexpected:
        return []

    column_indexes = {header: index + 1 for index, header in enumerate(headers)}
    parsed: list[ParsedRow] = []
    seen_keys: dict[tuple[Any, ...], int] = {}
    for row_number in range(2, sheet.max_row + 1):
        cells = [sheet.cell(row=row_number, column=index) for index in range(1, max(sheet.max_column, 1) + 1)]
        if all(_is_blank(cell.value) for cell in cells):
            continue
        row_has_error = False
        values: dict[str, Any] = {}
        for spec in schema.columns:
            cell = sheet.cell(row=row_number, column=column_indexes[spec.name])
            if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                _issue(
                    issues,
                    name,
                    row_number,
                    spec.name,
                    "FORMULA_NOT_ALLOWED",
                    "Formulas are not allowed; paste the computed value instead.",
                )
                row_has_error = True
                continue
            try:
                values[spec.name] = _normalize_cell(cell.value, spec)
            except ValueError as exc:
                _issue(issues, name, row_number, spec.name, "INVALID_VALUE", str(exc))
                row_has_error = True
        for cell in cells[len(headers) :]:
            if not _is_blank(cell.value):
                code = "FORMULA_NOT_ALLOWED" if cell.data_type == "f" else "UNEXPECTED_VALUE"
                _issue(
                    issues,
                    name,
                    row_number,
                    cell.coordinate,
                    code,
                    "Values outside the declared columns are not allowed.",
                )
                row_has_error = True
        if row_has_error:
            continue
        key = tuple(values.get(field) for field in schema.key)
        if key in seen_keys:
            _issue(
                issues,
                name,
                row_number,
                ",".join(schema.key),
                "DUPLICATE_KEY",
                f"Duplicate row key; first seen at row {seen_keys[key]}.",
            )
            continue
        seen_keys[key] = row_number
        parsed.append(ParsedRow(row_number, values))
    return parsed


def _row_index(rows: Mapping[str, list[ParsedRow]], sheet: str, fields: tuple[str, ...]) -> set[tuple[Any, ...]]:
    return {tuple(row.values.get(field) for field in fields) for row in rows.get(sheet, [])}


def _require_reference(
    rows: Mapping[str, list[ParsedRow]],
    issues: list[Issue],
    *,
    sheet: str,
    columns: tuple[str, ...],
    target_sheet: str,
    target_columns: tuple[str, ...],
    optional: bool = False,
) -> None:
    targets = _row_index(rows, target_sheet, target_columns)
    for row in rows.get(sheet, []):
        key = tuple(row.values.get(column) for column in columns)
        if optional and all(value is None for value in key):
            continue
        if key not in targets:
            _issue(
                issues,
                sheet,
                row.row_number,
                ",".join(columns),
                "UNKNOWN_REFERENCE",
                f"No matching {target_sheet} row exists for {key!r}.",
            )


def _validate_references(rows: Mapping[str, list[ParsedRow]], issues: list[Issue]) -> None:
    references = (
        ("Departments", ("college_code",), "Colleges", ("code",), False),
        ("Programs", ("department_code",), "Departments", ("code",), False),
        ("ProgramSubjects", ("program_code",), "Programs", ("code",), False),
        ("ProgramSubjects", ("subject_code",), "Subjects", ("code",), False),
        ("ProgramSubjects", ("authoritative_college_code",), "Colleges", ("code",), False),
        (
            "ProgramSubjects",
            ("authoritative_department_code",),
            "Departments",
            ("code",),
            True,
        ),
        ("Sections", ("program_code",), "Programs", ("code",), False),
        ("Instructors", ("department_code",), "Departments", ("code",), False),
        ("Rooms", ("owning_college_code",), "Colleges", ("code",), True),
        ("Rooms", ("owning_department_code",), "Departments", ("code",), True),
        ("RoomCapabilities", ("room_code",), "Rooms", ("code",), False),
        ("RoomCapabilities", ("capability_code",), "Capabilities", ("code",), False),
        ("RoomAuthorizations", ("room_code",), "Rooms", ("code",), False),
        ("RoomAuthorizations", ("college_code",), "Colleges", ("code",), True),
        ("RoomAuthorizations", ("department_code",), "Departments", ("code",), True),
        ("CourseOfferings", ("subject_code",), "Subjects", ("code",), False),
        (
            "CourseOfferings",
            ("offering_department_code",),
            "Departments",
            ("code",),
            False,
        ),
        ("OfferingSections", ("offering_key",), "CourseOfferings", ("external_key",), False),
        ("OfferingSections", ("section_code",), "Sections", ("code",), False),
        (
            "OfferingSections",
            ("program_code", "subject_code", "curriculum_version"),
            "ProgramSubjects",
            ("program_code", "subject_code", "curriculum_version"),
            False,
        ),
        ("OfferingInstructors", ("offering_key",), "CourseOfferings", ("external_key",), False),
        ("OfferingInstructors", ("instructor_code",), "Instructors", ("employee_code",), False),
        (
            "MeetingRequirements",
            ("offering_key",),
            "CourseOfferings",
            ("external_key",),
            False,
        ),
        (
            "MeetingCapabilities",
            ("meeting_key",),
            "MeetingRequirements",
            ("meeting_key",),
            False,
        ),
        ("MeetingCapabilities", ("capability_code",), "Capabilities", ("code",), False),
        (
            "InstructorAvailability",
            ("instructor_code",),
            "Instructors",
            ("employee_code",),
            False,
        ),
        ("RoomAvailability", ("room_code",), "Rooms", ("code",), False),
        ("Students", (), "Students", (), True),
        (
            "StudentSections",
            ("pseudonymous_code",),
            "Students",
            ("pseudonymous_code",),
            False,
        ),
        ("StudentSections", ("section_code",), "Sections", ("code",), False),
        (
            "InstructorPreferences",
            ("instructor_code",),
            "Instructors",
            ("employee_code",),
            False,
        ),
        ("LaboratoryProfiles", ("room_code",), "Rooms", ("code",), False),
        ("Locks", ("meeting_key",), "MeetingRequirements", ("meeting_key",), False),
        ("Locks", ("room_code",), "Rooms", ("code",), False),
    )
    for sheet, columns, target_sheet, target_columns, optional in references:
        if columns:
            _require_reference(
                rows,
                issues,
                sheet=sheet,
                columns=columns,
                target_sheet=target_sheet,
                target_columns=target_columns,
                optional=optional,
            )

    slot_keys = _row_index(rows, "TimeSlots", ("day", "sequence"))
    for sheet in ("InstructorAvailability", "RoomAvailability", "InstructorPreferences", "Locks"):
        for row in rows.get(sheet, []):
            key = (row.values["day"], row.values["sequence"])
            if key not in slot_keys:
                _issue(
                    issues,
                    sheet,
                    row.row_number,
                    "day,sequence",
                    "UNKNOWN_REFERENCE",
                    f"No TimeSlots row exists for {key!r}.",
                )


def _validate_semantics(rows: Mapping[str, list[ParsedRow]], term: models.AcademicTerm, issues: list[Issue]) -> None:
    minimum_rows = (
        "Colleges",
        "Departments",
        "Programs",
        "Subjects",
        "ProgramSubjects",
        "Sections",
        "Instructors",
        "Rooms",
        "TimeSlots",
        "CourseOfferings",
        "OfferingSections",
        "OfferingInstructors",
        "MeetingRequirements",
    )
    for sheet in minimum_rows:
        if not rows.get(sheet):
            _issue(
                issues,
                sheet,
                None,
                "",
                "INCOMPLETE_DATASET",
                f"At least one data row is required in {sheet}.",
            )

    department_colleges = {row.values["code"]: row.values["college_code"] for row in rows["Departments"]}
    sections = {row.values["code"]: row.values["program_code"] for row in rows["Sections"]}
    offerings = {row.values["external_key"]: row.values for row in rows["CourseOfferings"]}
    room_kinds = {row.values["code"]: row.values["kind"] for row in rows["Rooms"]}

    for row in rows["ProgramSubjects"]:
        department_code = row.values.get("authoritative_department_code")
        if department_code and department_colleges.get(department_code) != row.values["authoritative_college_code"]:
            _issue(
                issues,
                "ProgramSubjects",
                row.row_number,
                "authoritative_department_code",
                "UNIT_MISMATCH",
                "The authoritative department does not belong to the authoritative college.",
            )

    for row in rows["Rooms"]:
        has_college = bool(row.values.get("owning_college_code"))
        has_department = bool(row.values.get("owning_department_code"))
        if has_college == has_department:
            _issue(
                issues,
                "Rooms",
                row.row_number,
                "owning_college_code,owning_department_code",
                "INVALID_OWNER",
                "Provide exactly one college or department owner.",
            )
        if row.values["campus"].casefold() != term.campus.casefold():
            _issue(
                issues,
                "Rooms",
                row.row_number,
                "campus",
                "CAMPUS_MISMATCH",
                f"This one-campus import must use {term.campus!r}.",
            )

    for row in rows["RoomAuthorizations"]:
        if bool(row.values.get("college_code")) == bool(row.values.get("department_code")):
            _issue(
                issues,
                "RoomAuthorizations",
                row.row_number,
                "college_code,department_code",
                "INVALID_AUTHORIZATION_TARGET",
                "Provide exactly one authorized college or department.",
            )

    for row in rows["OfferingSections"]:
        section_program = sections.get(row.values["section_code"])
        offering = offerings.get(row.values["offering_key"])
        if section_program and section_program != row.values["program_code"]:
            _issue(
                issues,
                "OfferingSections",
                row.row_number,
                "program_code",
                "PROGRAM_MISMATCH",
                "The curriculum program does not match the section's program.",
            )
        if offering and offering["subject_code"] != row.values["subject_code"]:
            _issue(
                issues,
                "OfferingSections",
                row.row_number,
                "subject_code",
                "SUBJECT_MISMATCH",
                "The curriculum subject does not match the offered subject.",
            )

    slot_by_day: dict[int, list[ParsedRow]] = {}
    for row in rows["TimeSlots"]:
        start = datetime.strptime(row.values["starts_at"], "%H:%M")
        end = datetime.strptime(row.values["ends_at"], "%H:%M")
        if end <= start:
            _issue(
                issues,
                "TimeSlots",
                row.row_number,
                "ends_at",
                "INVALID_TIME_RANGE",
                "Slot end time must be after its start time.",
            )
        if row.values["sequence"] < 0:
            _issue(
                issues,
                "TimeSlots",
                row.row_number,
                "sequence",
                "INVALID_SEQUENCE",
                "Sequence cannot be negative.",
            )
        slot_by_day.setdefault(row.values["day"], []).append(row)
    for day_rows in slot_by_day.values():
        ordered = sorted(day_rows, key=lambda parsed: parsed.values["sequence"])
        for left, right in zip(ordered, ordered[1:], strict=False):
            left_end = datetime.strptime(left.values["ends_at"], "%H:%M")
            right_start = datetime.strptime(right.values["starts_at"], "%H:%M")
            if right_start < left_end:
                _issue(
                    issues,
                    "TimeSlots",
                    right.row_number,
                    "starts_at",
                    "OVERLAPPING_SLOT_DEFINITION",
                    f"This slot overlaps the slot at row {left.row_number}.",
                )

    offerings_with_sections = {row.values["offering_key"] for row in rows["OfferingSections"]}
    offerings_with_instructors = {row.values["offering_key"] for row in rows["OfferingInstructors"]}
    offerings_with_meetings = {row.values["offering_key"] for row in rows["MeetingRequirements"]}
    for row in rows["CourseOfferings"]:
        key = row.values["external_key"]
        if row.values.get("is_active") is False:
            continue
        for code, linked, label in (
            ("MISSING_OFFERING_SECTION", offerings_with_sections, "section"),
            ("MISSING_OFFERING_INSTRUCTOR", offerings_with_instructors, "instructor"),
            ("MISSING_MEETING_REQUIREMENT", offerings_with_meetings, "meeting requirement"),
        ):
            if key not in linked:
                _issue(
                    issues,
                    "CourseOfferings",
                    row.row_number,
                    "external_key",
                    code,
                    f"Active offering {key!r} has no {label} row.",
                )

    meeting_capabilities = {row.values["meeting_key"] for row in rows["MeetingCapabilities"]}
    for row in rows["MeetingRequirements"]:
        if row.values["occurrence_number"] < 1:
            _issue(
                issues,
                "MeetingRequirements",
                row.row_number,
                "occurrence_number",
                "INVALID_OCCURRENCE",
                "Occurrence number must be at least 1.",
            )
        if row.values["duration_atoms"] < 1:
            _issue(
                issues,
                "MeetingRequirements",
                row.row_number,
                "duration_atoms",
                "INVALID_DURATION",
                "Duration must be at least one atom.",
            )
        if row.values["component"] == models.MeetingComponent.LABORATORY and row.values["meeting_key"] not in meeting_capabilities:
            _issue(
                issues,
                "MeetingRequirements",
                row.row_number,
                "component",
                "LAB_CAPABILITY_REQUIRED",
                "Laboratory meetings must declare at least one required capability.",
            )

    instructor_availability = {row.values["instructor_code"] for row in rows["InstructorAvailability"]}
    for row in rows["Instructors"]:
        if row.values.get("is_active") is False:
            continue
        if not row.values.get("assume_fully_available") and row.values["employee_code"] not in instructor_availability:
            _issue(
                issues,
                "Instructors",
                row.row_number,
                "assume_fully_available",
                "MISSING_AVAILABILITY_PROFILE",
                "Provide availability rows or explicitly set assume_fully_available to TRUE.",
            )

    room_availability = {row.values["room_code"] for row in rows["RoomAvailability"]}
    for row in rows["Rooms"]:
        if row.values.get("is_active") is False:
            continue
        if not row.values.get("assume_fully_available") and row.values["code"] not in room_availability:
            _issue(
                issues,
                "Rooms",
                row.row_number,
                "assume_fully_available",
                "MISSING_AVAILABILITY_PROFILE",
                "Provide availability rows or explicitly set assume_fully_available to TRUE.",
            )

    laboratory_profiles = {row.values["room_code"] for row in rows["LaboratoryProfiles"]}
    for room_code, kind in room_kinds.items():
        if kind == models.RoomKind.LABORATORY and room_code not in laboratory_profiles:
            room_row = next(row for row in rows["Rooms"] if row.values["code"] == room_code)
            _issue(
                issues,
                "Rooms",
                room_row.row_number,
                "kind",
                "MISSING_LABORATORY_PROFILE",
                "Laboratory rooms require a LaboratoryProfiles row.",
            )


def _persist_issues(batch: models.ImportBatch, issues: list[Issue]) -> None:
    batch.errors.all().delete()
    models.ImportError.objects.bulk_create(
        [
            models.ImportError(
                batch=batch,
                sheet_name=issue.sheet_name,
                row_number=issue.row_number,
                column_name=issue.column_name,
                code=issue.code,
                message=issue.message,
            )
            for issue in issues
        ]
    )


@transaction.atomic
def preview_workbook(content: bytes, term: models.AcademicTerm, user: models.User) -> models.ImportBatch:
    """Parse and validate a workbook, persisting only normalized staging rows.

    The returned batch has ``PREVIEWED`` status only when it can be committed.
    Invalid workbooks still return an ``INVALID`` batch with cell-level errors.
    Re-previewing identical bytes returns and refreshes the existing uncommitted
    batch because the data model intentionally deduplicates files per term.
    """

    if not isinstance(content, bytes):
        raise TypeError("Workbook content must be bytes.")
    digest = hashlib.sha256(content).hexdigest()
    batch, created = models.ImportBatch.objects.get_or_create(
        term=term,
        file_hash=digest,
        defaults={
            "uploaded_by": user,
            "original_filename": f"semester-import-{digest[:12]}.xlsx",
        },
    )
    if not created and batch.status == models.ImportStatus.COMMITTED:
        return batch

    issues: list[Issue] = []
    parsed: dict[str, list[ParsedRow]] = {name: [] for name in SHEET_SCHEMAS}
    schema_version: str | None = None
    if not content:
        _issue(issues, "(workbook)", None, "", "EMPTY_WORKBOOK", "Workbook content is empty.")
    elif len(content) > MAX_WORKBOOK_BYTES:
        _issue(
            issues,
            "(workbook)",
            None,
            "",
            "WORKBOOK_TOO_LARGE",
            f"Workbook exceeds the {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB safety limit.",
        )
    else:
        try:
            workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
        except Exception as exc:  # openpyxl emits several format-specific exception classes
            _issue(
                issues,
                "(workbook)",
                None,
                "",
                "INVALID_WORKBOOK",
                f"The file is not a readable XLSX workbook: {exc}",
            )
        else:
            schema_version = _read_schema_version(workbook, issues)
            allowed_sheets = set(SHEET_SCHEMAS) | {SCHEMA_SHEET}
            for unexpected in sorted(set(workbook.sheetnames) - allowed_sheets):
                _issue(
                    issues,
                    unexpected,
                    None,
                    "",
                    "UNEXPECTED_SHEET",
                    f"Sheet {unexpected!r} is not part of schema {SCHEMA_VERSION}.",
                )
            for name, schema in SHEET_SCHEMAS.items():
                parsed[name] = _parse_sheet(workbook, name, schema, issues)
            _validate_references(parsed, issues)
            _validate_semantics(parsed, term, issues)

    normalized_sheets = {
        name: [row.values for row in parsed[name]]
        for name in SHEET_SCHEMAS
        if name in CORE_SHEETS or parsed[name]
    }
    batch.summary = {
        "schema_version": schema_version or SCHEMA_VERSION,
        "source_sha256": digest,
        "sheets": normalized_sheets,
        "row_counts": {name: len(rows) for name, rows in parsed.items()},
    }
    batch.total_rows = sum(len(rows) for rows in parsed.values())
    batch.error_count = len(issues)
    batch.status = models.ImportStatus.INVALID if issues else models.ImportStatus.PREVIEWED
    batch.uploaded_by = user
    batch.save(update_fields=[
        "summary",
        "total_rows",
        "error_count",
        "status",
        "uploaded_by",
        "updated_at",
    ])
    _persist_issues(batch, issues)
    models.AuditLog.objects.create(
        actor=user,
        action="import.previewed",
        entity_type="ImportBatch",
        entity_id=str(batch.pk),
        details={
            "term_id": term.pk,
            "file_hash": digest,
            "status": batch.status,
            "total_rows": batch.total_rows,
            "error_count": batch.error_count,
        },
    )
    return batch


def _rows(summary: Mapping[str, Any], sheet: str) -> list[dict[str, Any]]:
    sheets = summary.get("sheets")
    if not isinstance(sheets, dict) or not isinstance(sheets.get(sheet, []), list):
        raise ImportCommitError(f"Staged data for {sheet!r} is missing or malformed.")
    return sheets.get(sheet, [])


def _bool(row: Mapping[str, Any], key: str, default: bool = False) -> bool:
    value = row.get(key)
    return default if value is None else bool(value)


def _time(value: str) -> time:
    return datetime.strptime(value, "%H:%M").time()


def _validated_save(instance: models.Model) -> models.Model:
    instance.full_clean()
    instance.save()
    return instance


def _assert_commit_permission(user: models.User) -> None:
    if not user.is_active or (not user.is_superuser and user.role not in {
        models.UserRole.SYSTEM_ADMIN,
        models.UserRole.CENTRAL_SCHEDULER,
    }):
        raise ImportCommitError("Only a system administrator or central scheduler may commit imports.")


@transaction.atomic
def commit_import(batch: models.ImportBatch, user: models.User) -> models.TermDatasetRevision:
    """Atomically materialize one clean preview as a committed dataset revision."""

    _assert_commit_permission(user)
    # Do not join the nullable committed_revision relation while taking the
    # row lock. PostgreSQL rejects ``FOR UPDATE`` when it is applied to the
    # nullable side of an outer join. The foreign-key id below is sufficient
    # for the idempotency check, so only the non-null term relation is loaded.
    locked = models.ImportBatch.objects.select_for_update().select_related("term").get(pk=batch.pk)
    if locked.status == models.ImportStatus.COMMITTED or locked.committed_revision_id:
        raise ImportCommitError("This import batch has already been committed.")
    if locked.status != models.ImportStatus.PREVIEWED or locked.error_count or locked.errors.exists():
        raise ImportCommitError("Only a clean PREVIEWED import batch can be committed.")
    summary = locked.summary
    if not isinstance(summary, dict) or summary.get("schema_version") != SCHEMA_VERSION:
        raise ImportCommitError(f"The staged workbook does not use supported schema {SCHEMA_VERSION}.")
    if summary.get("source_sha256") != locked.file_hash:
        raise ImportCommitError("The staged workbook hash does not match its import batch.")

    try:
        next_revision = (
            models.TermDatasetRevision.objects.filter(term=locked.term).aggregate(value=Max("revision_number"))[
                "value"
            ]
            or 0
        ) + 1
        revision = models.TermDatasetRevision.objects.create(
            term=locked.term,
            revision_number=next_revision,
            status=models.RevisionStatus.DRAFT,
            label=f"Import {locked.file_hash[:12]}",
            created_by=user,
        )

        colleges: dict[str, models.College] = {}
        for row in _rows(summary, "Colleges"):
            college, _ = models.College.objects.update_or_create(
                code=row["code"],
                defaults={"name": row["name"], "is_active": _bool(row, "is_active", True)},
            )
            colleges[row["code"]] = college

        departments: dict[str, models.Department] = {}
        for row in _rows(summary, "Departments"):
            department, _ = models.Department.objects.update_or_create(
                code=row["code"],
                defaults={
                    "name": row["name"],
                    "college": colleges[row["college_code"]],
                    "is_active": _bool(row, "is_active", True),
                },
            )
            departments[row["code"]] = department

        programs: dict[str, models.Program] = {}
        for row in _rows(summary, "Programs"):
            program, _ = models.Program.objects.update_or_create(
                code=row["code"],
                defaults={
                    "name": row["name"],
                    "department": departments[row["department_code"]],
                    "curriculum_label": row.get("curriculum_label") or "",
                    "is_active": _bool(row, "is_active", True),
                },
            )
            programs[row["code"]] = program

        subjects: dict[str, models.Subject] = {}
        for row in _rows(summary, "Subjects"):
            subject, _ = models.Subject.objects.update_or_create(
                code=row["code"],
                defaults={
                    "title": row["title"],
                    "description": row.get("description") or "",
                    "is_active": _bool(row, "is_active", True),
                },
            )
            subjects[row["code"]] = subject

        program_subjects: dict[tuple[str, str, str], models.ProgramSubject] = {}
        for row in _rows(summary, "ProgramSubjects"):
            key = (row["program_code"], row["subject_code"], row["curriculum_version"])
            program_subject, _ = models.ProgramSubject.objects.update_or_create(
                program=programs[row["program_code"]],
                subject=subjects[row["subject_code"]],
                curriculum_version=row["curriculum_version"],
                defaults={
                    "classification": row["classification"],
                    "authoritative_college": colleges[row["authoritative_college_code"]],
                    "authoritative_department": departments.get(row.get("authoritative_department_code")),
                    "is_active": _bool(row, "is_active", True),
                },
            )
            program_subject.full_clean()
            program_subject.save()
            program_subjects[key] = program_subject

        instructors: dict[str, models.Instructor] = {}
        instructor_rows: dict[str, dict[str, Any]] = {}
        for row in _rows(summary, "Instructors"):
            instructor, _ = models.Instructor.objects.update_or_create(
                employee_code=row["employee_code"],
                defaults={
                    "display_name": row["display_name"],
                    "department": departments[row["department_code"]],
                    "is_active": _bool(row, "is_active", True),
                },
            )
            instructors[row["employee_code"]] = instructor
            instructor_rows[row["employee_code"]] = row

        rooms: dict[str, models.Room] = {}
        room_rows: dict[str, dict[str, Any]] = {}
        for row in _rows(summary, "Rooms"):
            room, _ = models.Room.objects.update_or_create(
                campus=row["campus"],
                code=row["code"],
                defaults={
                    "name": row.get("name") or "",
                    "kind": row["kind"],
                    "owning_college": colleges.get(row.get("owning_college_code")),
                    "owning_department": departments.get(row.get("owning_department_code")),
                    "is_active": _bool(row, "is_active", True),
                },
            )
            room.full_clean()
            room.save()
            rooms[row["code"]] = room
            room_rows[row["code"]] = row

        capabilities: dict[str, models.Capability] = {}
        for row in _rows(summary, "Capabilities"):
            capability, _ = models.Capability.objects.update_or_create(
                code=row["code"],
                defaults={"name": row["name"], "description": row.get("description") or ""},
            )
            capabilities[row["code"]] = capability

        for row in _rows(summary, "LaboratoryProfiles"):
            profile, _ = models.LaboratoryProfile.objects.update_or_create(
                room=rooms[row["room_code"]],
                defaults={
                    "laboratory_type": row["laboratory_type"],
                    "notes": row.get("notes") or "",
                },
            )
            profile.full_clean()
            profile.save()

        for row in _rows(summary, "RoomCapabilities"):
            models.RoomCapability.objects.get_or_create(
                room=rooms[row["room_code"]],
                capability=capabilities[row["capability_code"]],
            )

        sections: dict[str, models.Section] = {}
        for row in _rows(summary, "Sections"):
            section = models.Section(
                revision=revision,
                program=programs[row["program_code"]],
                code=row["code"],
                year_level=row["year_level"],
                cohort_status=row["cohort_status"],
                is_active=_bool(row, "is_active", True),
            )
            _validated_save(section)
            sections[row["code"]] = section

        for row in _rows(summary, "RoomAuthorizations"):
            authorization = models.RoomAuthorization(
                revision=revision,
                room=rooms[row["room_code"]],
                classification=row["classification"],
                college=colleges.get(row.get("college_code")),
                department=departments.get(row.get("department_code")),
                notes=row.get("notes") or "",
            )
            _validated_save(authorization)

        slots: dict[tuple[int, int], models.TimeSlot] = {}
        for row in _rows(summary, "TimeSlots"):
            slot = models.TimeSlot(
                revision=revision,
                day=row["day"],
                sequence=row["sequence"],
                starts_at=_time(row["starts_at"]),
                ends_at=_time(row["ends_at"]),
                is_break=_bool(row, "is_break"),
                is_active=_bool(row, "is_active", True),
            )
            _validated_save(slot)
            slots[(row["day"], row["sequence"])] = slot

        instructor_profiles: dict[str, models.InstructorAvailabilityProfile] = {}
        for code, instructor in instructors.items():
            assume_full = _bool(instructor_rows[code], "assume_fully_available")
            profile = models.InstructorAvailabilityProfile(
                revision=revision,
                instructor=instructor,
                assume_fully_available=assume_full,
                acknowledged_by=user if assume_full else None,
                acknowledged_at=timezone.now() if assume_full else None,
            )
            _validated_save(profile)
            instructor_profiles[code] = profile
        for row in _rows(summary, "InstructorAvailability"):
            availability = models.InstructorAvailability(
                profile=instructor_profiles[row["instructor_code"]],
                time_slot=slots[(row["day"], row["sequence"])],
                is_available=row["is_available"],
            )
            _validated_save(availability)

        room_profiles: dict[str, models.RoomAvailabilityProfile] = {}
        for code, room in rooms.items():
            assume_full = _bool(room_rows[code], "assume_fully_available")
            profile = models.RoomAvailabilityProfile(
                revision=revision,
                room=room,
                assume_fully_available=assume_full,
                acknowledged_by=user if assume_full else None,
                acknowledged_at=timezone.now() if assume_full else None,
            )
            _validated_save(profile)
            room_profiles[code] = profile
        for row in _rows(summary, "RoomAvailability"):
            availability = models.RoomAvailability(
                profile=room_profiles[row["room_code"]],
                time_slot=slots[(row["day"], row["sequence"])],
                is_available=row["is_available"],
            )
            _validated_save(availability)

        offerings: dict[str, models.CourseOffering] = {}
        for row in _rows(summary, "CourseOfferings"):
            offering = models.CourseOffering(
                revision=revision,
                subject=subjects[row["subject_code"]],
                offering_department=departments[row["offering_department_code"]],
                external_key=row["external_key"],
                is_active=_bool(row, "is_active", True),
            )
            _validated_save(offering)
            offerings[row["external_key"]] = offering

        for row in _rows(summary, "OfferingSections"):
            link = models.OfferingSection(
                offering=offerings[row["offering_key"]],
                section=sections[row["section_code"]],
                program_subject=program_subjects[
                    (row["program_code"], row["subject_code"], row["curriculum_version"])
                ],
            )
            _validated_save(link)
        for row in _rows(summary, "OfferingInstructors"):
            link = models.OfferingInstructor(
                offering=offerings[row["offering_key"]],
                instructor=instructors[row["instructor_code"]],
            )
            _validated_save(link)

        meetings: dict[str, models.MeetingRequirement] = {}
        for row in _rows(summary, "MeetingRequirements"):
            stable_key = uuid.uuid5(
                uuid.NAMESPACE_URL,
                f"usm-scheduler:{locked.term_id}:{revision.revision_number}:{row['meeting_key']}",
            )
            meeting = models.MeetingRequirement(
                offering=offerings[row["offering_key"]],
                stable_key=stable_key,
                component=row["component"],
                occurrence_number=row["occurrence_number"],
                duration_atoms=row["duration_atoms"],
                distinct_day_group=row.get("distinct_day_group") or "",
                is_active=_bool(row, "is_active", True),
            )
            _validated_save(meeting)
            meetings[row["meeting_key"]] = meeting
        for row in _rows(summary, "MeetingCapabilities"):
            link = models.MeetingRequiredCapability(
                meeting_requirement=meetings[row["meeting_key"]],
                capability=capabilities[row["capability_code"]],
            )
            _validated_save(link)

        for row in _rows(summary, "InstructorPreferences"):
            preference = models.InstructorPreference(
                profile=instructor_profiles[row["instructor_code"]],
                time_slot=slots[(row["day"], row["sequence"])],
                level=row["level"],
                weight=row["weight"],
            )
            _validated_save(preference)

        students: dict[str, models.Student] = {}
        for row in _rows(summary, "Students"):
            student, _ = models.Student.objects.update_or_create(
                pseudonymous_code=row["pseudonymous_code"],
                defaults={"status": row["status"]},
            )
            students[row["pseudonymous_code"]] = student
        for row in _rows(summary, "StudentSections"):
            models.StudentSectionMembership.objects.create(
                student=students[row["pseudonymous_code"]],
                section=sections[row["section_code"]],
            )

        for row in _rows(summary, "Locks"):
            lock = models.LockedAssignment(
                meeting_requirement=meetings[row["meeting_key"]],
                room=rooms[row["room_code"]],
                start_time_slot=slots[(row["day"], row["sequence"])],
                locked_by=user,
                reason=row["reason"],
            )
            _validated_save(lock)

        revision.content_hash = models.canonical_sha256(
            {
                "schema_version": SCHEMA_VERSION,
                "term_id": locked.term_id,
                "sheets": summary["sheets"],
            }
        )
        revision.status = models.RevisionStatus.COMMITTED
        revision.save()
        locked.status = models.ImportStatus.COMMITTED
        locked.committed_revision = revision
        locked.save(update_fields=["status", "committed_revision", "updated_at"])
        models.AuditLog.objects.create(
            actor=user,
            action="import.committed",
            entity_type="TermDatasetRevision",
            entity_id=str(revision.pk),
            details={
                "import_batch_id": locked.pk,
                "term_id": locked.term_id,
                "revision_number": revision.revision_number,
                "content_hash": revision.content_hash,
            },
        )
        return revision
    except ImportCommitError:
        raise
    except (KeyError, TypeError, ValueError, ValidationError, IntegrityError) as exc:
        raise ImportCommitError(f"Import commit failed atomically: {exc}") from exc


def build_import_template() -> bytes:
    """Return a styled, header-only XLSX template for the current schema."""

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    schema_sheet = workbook.create_sheet(SCHEMA_SHEET)
    schema_sheet["A1"] = "schema_version"
    schema_sheet["B1"] = SCHEMA_VERSION
    schema_sheet["A2"] = "day_numbering"
    schema_sheet["B2"] = "0=Monday ... 6=Sunday"
    schema_sheet.sheet_state = "hidden"

    header_fill = PatternFill(fill_type="solid", fgColor="073A2A")
    header_font = Font(color="FFFFFF", bold=True)
    for name, schema in SHEET_SCHEMAS.items():
        sheet = workbook.create_sheet(name)
        sheet.freeze_panes = "A2"
        for column_number, spec in enumerate(schema.columns, start=1):
            cell = sheet.cell(row=1, column=column_number, value=spec.name)
            cell.fill = header_fill
            cell.font = header_font
            cell.comment = Comment(
                (spec.description + " " if spec.description else "")
                + ("Required." if spec.required else "Optional; blank uses the documented default."),
                "USM Scheduler",
            )
            sheet.column_dimensions[cell.column_letter].width = max(14, min(35, len(spec.name) + 4))
            if spec.choices:
                formula = '"' + ",".join(spec.choices) + '"'
                validation = DataValidation(type="list", formula1=formula, allow_blank=not spec.required)
                validation.error = f"Choose one of: {', '.join(spec.choices)}"
                validation.errorTitle = "Invalid value"
                sheet.add_data_validation(validation)
                validation.add(f"{cell.column_letter}2:{cell.column_letter}10000")
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(schema.columns)).coordinate}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


create_import_template = build_import_template


__all__ = [
    "CORE_SHEETS",
    "OPTIONAL_SHEETS",
    "SCHEMA_VERSION",
    "SHEET_SCHEMAS",
    "ImportCommitError",
    "ImportPipelineError",
    "build_import_template",
    "commit_import",
    "create_import_template",
    "preview_workbook",
]
