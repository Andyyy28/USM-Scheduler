from __future__ import annotations

import hashlib
from datetime import date, datetime
from io import BytesIO

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from scheduler import models
from scheduler.domain import SolverAlgorithm, SolverConfig, SolverStatus
from scheduler.services.imports import commit_import, preview_workbook
from scheduler.services.problem_builder import build_problem
from scheduler.services.trial_data import (
    TRIAL_DAILY_LOAD_RULE_CODE,
    TRIAL_FIXED_RULE_CODE,
    TRIAL_RESERVED_BLOCK_RULE_CODE,
    approved_trial_policy_hashes,
    build_trial_workbook_bytes,
)
from scheduler.solvers import CpSatSolver, GeneticAlgorithmSolver, is_ortools_available

pytestmark = pytest.mark.django_db


def _term() -> models.AcademicTerm:
    return models.AcademicTerm.objects.create(
        academic_year="2026-2027",
        semester=models.Semester.FIRST,
        campus="Kabacan",
        starts_on=date(2026, 8, 1),
        ends_on=date(2026, 12, 20),
    )


def _approved_trial_policies(
    term: models.AcademicTerm,
    approver: models.User,
) -> dict[str, str]:
    approved_at = timezone.make_aware(datetime(2026, 7, 1, 8, 0))
    definitions = {
        TRIAL_FIXED_RULE_CODE: (
            "Fixed 50-student rule",
            {"fixed_student_limit": 50},
        ),
        TRIAL_DAILY_LOAD_RULE_CODE: (
            "Instructor daily teaching-atom limit",
            {"unit": "teaching_atom"},
        ),
        TRIAL_RESERVED_BLOCK_RULE_CODE: (
            "Recurring reserved teaching blocks",
            {"applies_to": "recurring_reserved_blocks"},
        ),
    }
    for code, (definition, parameters) in definitions.items():
        models.ConstraintPolicyVersion.objects.create(
            rule_code=code,
            version=1,
            title=definition,
            definition=definition,
            classification=models.ConstraintKind.HARD,
            owner_office="Synthetic test office",
            source="Synthetic test policy; not institutional evidence",
            effective_term=term,
            parameters=parameters,
            is_approved=True,
            approved_by=approver,
            approved_at=approved_at,
        )
    return approved_trial_policy_hashes(term)


def test_trial_download_is_deterministic_synthetic_and_role_restricted() -> None:
    central = models.User.objects.create_user(
        username="trial-central",
        role=models.UserRole.CENTRAL_SCHEDULER,
    )
    term = _term()
    policy_hashes = _approved_trial_policies(term, central)
    first = build_trial_workbook_bytes(**policy_hashes)
    second = build_trial_workbook_bytes(**policy_hashes)
    assert first == second
    assert hashlib.sha256(first).digest() == hashlib.sha256(second).digest()

    workbook = load_workbook(BytesIO(first), read_only=True)
    assert "SYNTHETIC TEST DATA ONLY" in workbook.properties.description
    assert workbook["Students"].max_row == 1
    assert workbook["MeetingRequirements"].max_row == 15
    assert workbook["_Schema"]["B1"].value == "1.1"
    assert workbook["_Schema"]["A3"].value == "fixed_student_limit"
    assert "1-50 students" in workbook["_Schema"]["B3"].value
    assert workbook["Sections"]["E2"].value == 25
    assert workbook["Sections"]["E4"].value == 50
    assert workbook["ConstraintPolicyReferences"].max_row == 4
    assert workbook["InstructorProfiles"].max_row == 7
    assert workbook["ReservedBlocks"].max_row == 2
    room_headers = {cell.value for cell in workbook["Rooms"][1]}
    assert "capacity" not in room_headers
    assert "chair_count" not in room_headers

    reviewer = models.User.objects.create_user(
        username="trial-reviewer",
        role=models.UserRole.COLLEGE_REVIEWER,
    )
    client = APIClient()
    client.force_authenticate(reviewer)
    assert client.get(reverse("api:trial-workbook"), {"term_id": term.pk}).status_code == 403

    client.force_authenticate(central)
    assert client.get(reverse("api:trial-workbook")).status_code == 400
    response = client.get(reverse("api:trial-workbook"), {"term_id": term.pk})
    assert response.status_code == 200
    assert response.content == first
    assert "Synthetic-Trial" in response["Content-Disposition"]
    assert response["X-USM-Target-Term-ID"] == str(term.pk)


def test_trial_download_requires_existing_approved_policies_without_creating_them() -> None:
    central = models.User.objects.create_user(
        username="trial-no-policy", role=models.UserRole.CENTRAL_SCHEDULER
    )
    term = _term()
    client = APIClient()
    client.force_authenticate(central)
    response = client.get(reverse("api:trial-workbook"), {"term_id": term.pk})
    assert response.status_code == 400
    assert "missing approved" in response.json()["term_id"]
    assert not models.ConstraintPolicyVersion.objects.exists()
    assert client.get(reverse("api:trial-workbook"), {"term_id": "invalid"}).status_code == 400


def test_trial_workbook_imports_and_preflights_deterministically() -> None:
    central = models.User.objects.create_user(
        username="trial-scheduler",
        role=models.UserRole.CENTRAL_SCHEDULER,
    )
    term = _term()
    policy_hashes = _approved_trial_policies(term, central)
    batch = preview_workbook(build_trial_workbook_bytes(**policy_hashes), term, central)

    assert batch.status == models.ImportStatus.PREVIEWED
    assert batch.data_origin == models.DatasetOrigin.SYNTHETIC
    assert batch.error_count == 0
    revision = commit_import(batch, central)
    assert revision.sections.count() == 5
    assert revision.course_offerings.count() == 11
    assert models.Student.objects.count() == 0
    assert models.LockedAssignment.objects.count() == 1
    assert sorted(revision.sections.values_list("expected_enrollment", flat=True)) == [25, 25, 30, 49, 50]
    assert revision.reserved_time_blocks.count() == 1
    assert revision.instructor_availability_profiles.filter(
        max_daily_teaching_atoms__isnull=False
    ).count() == 6

    objective = models.ObjectiveProfile.objects.create(
        name="Synthetic trial objective",
        term=revision.term,
        is_approved=True,
        approved_by=central,
    )
    result = build_problem(revision, objective)
    problem = result.problem
    assert len(problem.events) == 14
    assert result.candidate_count > 100
    assert len(problem.locked_assignments) == 1
    assert any(event.requires_laboratory_room for event in problem.events)
    assert any(len(event.section_ids) == 2 for event in problem.events)
    assert any(len(event.instructor_ids) == 2 for event in problem.events)
    assert all(event.candidates for event in problem.events)
    assert {
        lock.event_id: lock.candidate_id for lock in problem.locked_assignments
    }.items() <= {
        event.event_id: event.candidates[0].candidate_id for event in problem.events
    }.items()


def test_trial_workbook_rejects_conflicting_institutional_origin() -> None:
    central = models.User.objects.create_user(
        username="trial-origin-scheduler",
        role=models.UserRole.CENTRAL_SCHEDULER,
    )
    term = _term()
    policy_hashes = _approved_trial_policies(term, central)

    batch = preview_workbook(
        build_trial_workbook_bytes(**policy_hashes),
        term,
        central,
        data_origin=models.DatasetOrigin.INSTITUTIONAL,
    )

    assert batch.status == models.ImportStatus.INVALID
    assert batch.data_origin == models.DatasetOrigin.SYNTHETIC
    assert batch.errors.filter(code="DATA_ORIGIN_MISMATCH").exists()


@pytest.mark.diagnostic
@pytest.mark.skipif(not is_ortools_available(), reason="OR-Tools is not installed")
def test_trial_workbook_performance_exercise_is_feasible_for_both_engines() -> None:
    central = models.User.objects.create_user(
        username="trial-diagnostic-scheduler",
        role=models.UserRole.CENTRAL_SCHEDULER,
    )
    term = _term()
    policy_hashes = _approved_trial_policies(term, central)
    batch = preview_workbook(build_trial_workbook_bytes(**policy_hashes), term, central)
    revision = commit_import(batch, central)
    objective = models.ObjectiveProfile.objects.create(
        name="Synthetic trial diagnostic objective",
        term=revision.term,
        is_approved=True,
        approved_by=central,
    )
    problem = build_problem(revision, objective).problem

    cp_result = CpSatSolver().solve(
        problem,
        SolverConfig(
            algorithm=SolverAlgorithm.CP_SAT,
            seed=1001,
            time_limit_seconds=30,
            worker_count=1,
        ),
    )
    assert cp_result.status in {SolverStatus.FEASIBLE, SolverStatus.OPTIMAL}
    assert cp_result.validation.feasible

    ga_result = GeneticAlgorithmSolver().solve(
        problem,
        SolverConfig(
            algorithm=SolverAlgorithm.GENETIC_ALGORITHM,
            seed=1001,
            time_limit_seconds=30,
            worker_count=1,
            population_size=100,
            tournament_size=3,
            max_generations=100,
            repair_attempts=20,
        ),
    )
    assert ga_result.status == SolverStatus.FEASIBLE
    assert ga_result.validation.feasible
