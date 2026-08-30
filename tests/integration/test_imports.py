from __future__ import annotations

from io import BytesIO

import pytest
from django.core.management import call_command
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from scheduler import models
from scheduler.services.imports import (
    CORE_SHEETS,
    OPTIONAL_SHEETS,
    SCHEMA_SHEET,
    SCHEMA_VERSION,
    SHEET_SCHEMAS,
    ImportCommitError,
    build_import_template,
    commit_import,
    preview_workbook,
)
from scheduler.services.problem_builder import ProblemBuildError, build_problem

pytestmark = pytest.mark.django_db


def make_user(*, role: str = models.UserRole.CENTRAL_SCHEDULER) -> models.User:
    return models.User.objects.create_user(username=f"user-{role}", role=role)


def make_term() -> models.AcademicTerm:
    from datetime import date

    return models.AcademicTerm.objects.create(
        academic_year="2026-2027",
        semester=models.Semester.FIRST,
        campus="Kabacan",
        starts_on=date(2026, 8, 1),
        ends_on=date(2026, 12, 20),
    )


def workbook_bytes(workbook) -> bytes:  # type: ignore[no-untyped-def]
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_reviewer_import_page_does_not_disclose_batch_metadata(client) -> None:  # type: ignore[no-untyped-def]
    term = make_term()
    central = make_user()
    reviewer = models.User.objects.create_user(
        username="scoped-reviewer",
        role=models.UserRole.COLLEGE_REVIEWER,
    )
    models.ImportBatch.objects.create(
        term=term,
        uploaded_by=central,
        original_filename="confidential-usm-semester.xlsx",
        file_hash="b" * 64,
    )
    client.force_login(reviewer)

    response = client.get(reverse("scheduler:imports"))

    assert response.status_code == 200
    assert b"confidential-usm-semester.xlsx" not in response.content


def make_policy(
    term: models.AcademicTerm,
    user: models.User,
    *,
    rule_code: str = "APPROVED_SCHEDULING_RULES",
) -> models.ConstraintPolicyVersion:
    return models.ConstraintPolicyVersion.objects.create(
        rule_code=rule_code,
        version=1,
        title="Approved scheduling constraints",
        definition="Approved hard scheduling constraints used by the imported term.",
        classification=models.ConstraintKind.HARD,
        owner_office="Office of the University Registrar",
        source="Approved test policy",
        effective_term=term,
        parameters={"fixed_student_limit": 50},
        is_approved=True,
        approved_by=user,
        approved_at=timezone.now(),
    )


def valid_workbook(policy: models.ConstraintPolicyVersion):  # type: ignore[no-untyped-def]
    workbook = load_workbook(BytesIO(build_import_template()))
    rows = {
        "Colleges": ["CSM", "College of Science and Mathematics", True],
        "Departments": ["DCS", "Department of Computer Science", "CSM", True],
        "Programs": ["BSCS", "BS Computer Science", "DCS", "2026", True],
        "Subjects": ["CS101", "Introduction to Computing", "", True],
        "ProgramSubjects": ["BSCS", "CS101", "2026", "MAJOR", "CSM", "DCS", True],
        "Sections": ["BSCS-1A", "BSCS", 1, "INCOMING", 50, True],
        "Instructors": ["FAC-001", "Faculty One", "DCS", True, True],
        "Rooms": ["CLAB-1", "Computer Laboratory 1", "Kabacan", "LABORATORY", "", "DCS", True, True],
        "Capabilities": ["COMPUTER_LAB", "Computer laboratory", "Computers and projector"],
        "RoomCapabilities": ["CLAB-1", "COMPUTER_LAB"],
        "RoomAuthorizations": ["CLAB-1", "MAJOR", "", "DCS", "DCS major room"],
        "TimeSlots": [0, 0, "08:00", "08:30", False, True],
        "CourseOfferings": ["CS101-BSCS1A", "CS101", "DCS", True],
        "OfferingSections": ["CS101-BSCS1A", "BSCS-1A", "BSCS", "CS101", "2026"],
        "OfferingInstructors": ["CS101-BSCS1A", "FAC-001"],
        "MeetingRequirements": ["CS101-LAB-1", "CS101-BSCS1A", "LAB", 1, 2, "", True],
        "MeetingCapabilities": ["CS101-LAB-1", "COMPUTER_LAB"],
        "Students": ["anon-001", "ACTIVE"],
        "StudentSections": ["anon-001", "BSCS-1A"],
        "InstructorPreferences": ["FAC-001", 0, 0, "PREFERRED", 2],
        "ConstraintPolicyReferences": [policy.rule_code, policy.version, policy.policy_hash],
        "InstructorProfiles": ["FAC-001", 6, False, policy.rule_code, policy.version],
        "LaboratoryProfiles": ["CLAB-1", "Computer", "Teaching laboratory"],
        "Locks": ["CS101-LAB-1", "CLAB-1", 0, 0, "Existing approved placement"],
    }
    for sheet, row in rows.items():
        workbook[sheet].append(row)
    workbook["TimeSlots"].append([0, 1, "08:30", "09:00", False, True])
    return workbook


def test_template_is_versioned_and_contains_all_declared_sheets() -> None:
    content = build_import_template()
    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content))
    assert workbook[SCHEMA_SHEET]["A1"].value == "schema_version"
    assert workbook[SCHEMA_SHEET]["B1"].value == SCHEMA_VERSION
    assert set(CORE_SHEETS + OPTIONAL_SHEETS).issubset(workbook.sheetnames)
    for name, schema in SHEET_SCHEMAS.items():
        headers = [cell.value for cell in workbook[name][1]]
        assert headers == [column.name for column in schema.columns]
    room_headers = {str(cell.value).casefold() for cell in workbook["Rooms"][1]}
    assert not room_headers.intersection(
        {
            "capacity",
            "room_capacity",
            "chair_count",
            "seat_count",
            "floor_space",
            "floor_area",
            "physical_dimensions",
        }
    )


@pytest.mark.parametrize("expected_enrollment", [1, 49, 50])
def test_section_enrollment_accepts_fixed_rule_boundaries(expected_enrollment: int) -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    enrollment_column = [cell.value for cell in workbook["Sections"][1]].index(
        "expected_enrollment"
    ) + 1
    workbook["Sections"].cell(2, enrollment_column, expected_enrollment)

    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.PREVIEWED
    assert batch.summary["sheets"]["Sections"][0]["expected_enrollment"] == expected_enrollment


@pytest.mark.parametrize("expected_enrollment", [0, 51, 75])
def test_section_enrollment_rejects_values_outside_fixed_rule(
    expected_enrollment: int,
) -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    enrollment_column = [cell.value for cell in workbook["Sections"][1]].index(
        "expected_enrollment"
    ) + 1
    workbook["Sections"].cell(2, enrollment_column, expected_enrollment)

    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.INVALID
    issue = batch.errors.get(sheet_name="Sections", column_name="expected_enrollment")
    assert issue.code == "INVALID_VALUE"
    assert "1 to 50" in issue.message


def test_combined_meeting_total_of_exactly_50_builds_a_snapshot_problem() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    section_headers = [cell.value for cell in workbook["Sections"][1]]
    enrollment_column = section_headers.index("expected_enrollment") + 1
    workbook["Sections"].cell(2, enrollment_column, 20)
    workbook["Sections"].append(["BSCS-1B", "BSCS", 1, "INCOMING", 30, True])
    workbook["OfferingSections"].append(
        ["CS101-BSCS1A", "BSCS-1B", "BSCS", "CS101", "2026"]
    )
    revision = commit_import(
        preview_workbook(workbook_bytes(workbook), term, user),
        user,
    )
    objective = models.ObjectiveProfile.objects.create(
        name="Combined enrollment boundary objective",
        term=term,
    )

    problem = build_problem(revision, objective).problem

    assert problem.events[0].meeting_headcount == 50
    assert problem.events[0].section_headcounts == (
        (str(revision.sections.get(code="BSCS-1A").pk), 20),
        (str(revision.sections.get(code="BSCS-1B").pk), 30),
    )


def test_combined_meeting_total_above_50_blocks_problem_snapshot_preflight() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    workbook["Sections"].append(["BSCS-1B", "BSCS", 1, "INCOMING", 1, True])
    workbook["OfferingSections"].append(
        ["CS101-BSCS1A", "BSCS-1B", "BSCS", "CS101", "2026"]
    )
    revision = commit_import(
        preview_workbook(workbook_bytes(workbook), term, user),
        user,
    )
    objective = models.ObjectiveProfile.objects.create(
        name="Combined enrollment overflow objective",
        term=term,
    )

    with pytest.raises(ProblemBuildError) as captured:
        build_problem(revision, objective)

    issue = next(
        item
        for item in captured.value.issues
        if item.code == "MEETING_HEADCOUNT_EXCEEDS_FIXED_LIMIT"
    )
    assert "51 students" in issue.message
    assert "BSCS-1A=50" in issue.message
    assert "BSCS-1B=1" in issue.message
    assert issue.entity_type == "MeetingRequirement"


def test_instructor_profile_accepts_explicit_approved_no_limit_policy() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    headers = [cell.value for cell in workbook["InstructorProfiles"][1]]
    workbook["InstructorProfiles"].cell(
        2,
        headers.index("max_daily_teaching_atoms") + 1,
    ).value = None
    workbook["InstructorProfiles"].cell(
        2,
        headers.index("acknowledge_no_daily_limit") + 1,
        True,
    )

    batch = preview_workbook(workbook_bytes(workbook), term, user)
    revision = commit_import(batch, user)

    profile = revision.instructor_availability_profiles.get()
    assert profile.max_daily_teaching_atoms is None
    assert profile.acknowledge_no_daily_limit is True
    assert profile.daily_load_policy_version == policy


@pytest.mark.parametrize(
    ("maximum", "no_limit", "expected_code"),
    [
        (None, False, "MISSING_DAILY_LOAD_POLICY"),
        (6, True, "CONFLICTING_DAILY_LOAD_POLICY"),
        (0, False, "INVALID_VALUE"),
    ],
)
def test_instructor_profile_requires_exactly_one_valid_daily_load_policy(
    maximum: int | None,
    no_limit: bool,
    expected_code: str,
) -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    headers = [cell.value for cell in workbook["InstructorProfiles"][1]]
    workbook["InstructorProfiles"].cell(
        2,
        headers.index("max_daily_teaching_atoms") + 1,
    ).value = maximum
    workbook["InstructorProfiles"].cell(
        2,
        headers.index("acknowledge_no_daily_limit") + 1,
        no_limit,
    )

    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.INVALID
    assert expected_code in set(batch.errors.values_list("code", flat=True))


def test_policy_reference_requires_matching_approved_term_hash() -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    workbook["ConstraintPolicyReferences"]["C2"] = "f" * 64

    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.INVALID
    assert batch.errors.filter(code="POLICY_HASH_MISMATCH").exists()


def test_all_reserved_block_scopes_are_imported_with_slots_and_policy() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    scoped_rows = (
        ("institution", "INSTITUTION", None),
        ("college", "COLLEGE", "CSM"),
        ("department", "DEPARTMENT", "DCS"),
        ("program", "PROGRAM", "BSCS"),
        ("section", "SECTION", "BSCS-1A"),
    )
    for block_key, scope, target_code in scoped_rows:
        workbook["ReservedBlocks"].append(
            [
                block_key,
                scope,
                target_code,
                policy.rule_code,
                policy.version,
                f"Reserved {scope.lower()} block",
                "Approved recurring teaching block",
                True,
            ]
        )
        workbook["ReservedBlockSlots"].append([block_key, 0, 0])

    batch = preview_workbook(workbook_bytes(workbook), term, user)
    revision = commit_import(batch, user)

    assert batch.status == models.ImportStatus.COMMITTED
    assert revision.reserved_time_blocks.count() == len(scoped_rows)
    for block in revision.reserved_time_blocks.all():
        assert block.policy_version == policy
        assert block.slot_links.count() == 1
        if block.scope == models.ReservedBlockScope.INSTITUTION:
            assert block.scope_target is None
        else:
            assert block.scope_target is not None


@pytest.mark.parametrize(
    ("scope", "target_code"),
    [
        ("INSTITUTION", None),
        ("COLLEGE", "CSM"),
        ("DEPARTMENT", "DCS"),
        ("PROGRAM", "BSCS"),
        ("SECTION", "BSCS-1A"),
    ],
)
def test_each_reserved_block_scope_removes_the_same_meeting_domain(
    scope: str,
    target_code: str | None,
) -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    workbook["ReservedBlocks"].append(
        [
            "blocked-teaching-period",
            scope,
            target_code,
            policy.rule_code,
            policy.version,
            f"Reserved {scope.lower()} block",
            "Approved recurring teaching block",
            True,
        ]
    )
    workbook["ReservedBlockSlots"].append(["blocked-teaching-period", 0, 0])
    revision = commit_import(
        preview_workbook(workbook_bytes(workbook), term, user),
        user,
    )
    objective = models.ObjectiveProfile.objects.create(
        name=f"{scope} reserved-block test objective",
        term=term,
    )

    with pytest.raises(ProblemBuildError) as captured:
        build_problem(revision, objective)

    assert "EMPTY_CANDIDATE_DOMAIN" in {issue.code for issue in captured.value.issues}


def test_reserved_block_requires_matching_target_and_at_least_one_slot() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    workbook = valid_workbook(policy)
    workbook["ReservedBlocks"].append(
        [
            "bad-college-block",
            "COLLEGE",
            "UNKNOWN",
            policy.rule_code,
            policy.version,
            "Invalid block",
            "",
            True,
        ]
    )

    batch = preview_workbook(workbook_bytes(workbook), term, user)

    codes = set(batch.errors.values_list("code", flat=True))
    assert batch.status == models.ImportStatus.INVALID
    assert {"UNKNOWN_SCOPE_TARGET", "MISSING_RESERVED_BLOCK_SLOT"}.issubset(codes)


def test_schema_1_0_workbook_remains_importable_as_legacy_exploratory_data() -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    workbook[SCHEMA_SHEET]["B1"] = "1.0"
    for sheet_name in (
        "ConstraintPolicyReferences",
        "InstructorProfiles",
        "ReservedBlocks",
        "ReservedBlockSlots",
    ):
        del workbook[sheet_name]
    section_headers = [cell.value for cell in workbook["Sections"][1]]
    workbook["Sections"].delete_cols(section_headers.index("expected_enrollment") + 1)

    batch = preview_workbook(workbook_bytes(workbook), term, user)
    revision = commit_import(batch, user)

    assert batch.status == models.ImportStatus.COMMITTED
    assert batch.summary["schema_version"] == "1.0"
    assert revision.sections.get().expected_enrollment is None
    profile = revision.instructor_availability_profiles.get()
    assert profile.max_daily_teaching_atoms is None
    assert profile.acknowledge_no_daily_limit is False
    assert profile.daily_load_policy_version is None


def test_preview_stores_only_normalized_rows_and_commit_builds_complete_revision() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    batch = preview_workbook(workbook_bytes(valid_workbook(policy)), term, user)

    assert batch.status == models.ImportStatus.PREVIEWED
    assert batch.error_count == 0
    assert not batch.errors.exists()
    assert batch.summary["schema_version"] == SCHEMA_VERSION
    assert batch.summary["sheets"]["TimeSlots"][0]["starts_at"] == "08:00"
    assert batch.summary["sheets"]["Instructors"][0]["assume_fully_available"] is True
    assert batch.summary["sheets"]["Sections"][0]["expected_enrollment"] == 50
    assert "workbook" not in batch.summary

    revision = commit_import(batch, user)
    batch.refresh_from_db()
    assert revision.status == models.RevisionStatus.COMMITTED
    assert len(revision.content_hash) == 64
    assert batch.status == models.ImportStatus.COMMITTED
    assert batch.committed_revision == revision
    assert revision.sections.get().code == "BSCS-1A"
    assert revision.sections.get().expected_enrollment == 50
    assert revision.time_slots.count() == 2
    assert revision.course_offerings.count() == 1
    assert revision.course_offerings.get().sections.get().code == "BSCS-1A"
    assert revision.course_offerings.get().instructors.get().employee_code == "FAC-001"
    meeting = revision.course_offerings.get().meeting_requirements.get()
    assert meeting.duration_atoms == 2
    assert meeting.required_capabilities.get().code == "COMPUTER_LAB"
    assert models.LaboratoryProfile.objects.get().room.code == "CLAB-1"
    assert models.Student.objects.get().pseudonymous_code == "anon-001"
    assert models.InstructorPreference.objects.get().weight == 2
    assert models.LockedAssignment.objects.get().meeting_requirement == meeting
    assert revision.instructor_availability_profiles.get().assume_fully_available is True
    assert revision.instructor_availability_profiles.get().max_daily_teaching_atoms == 6
    assert revision.instructor_availability_profiles.get().daily_load_policy_version == policy
    assert revision.room_availability_profiles.get().assume_fully_available is True
    objective = models.ObjectiveProfile.objects.create(name="Imported term objective", term=term)
    problem = build_problem(revision, objective).problem
    assert len(problem.events) == 1
    assert len(problem.events[0].candidates) == 1
    assert problem.locked_assignments[0].event_id == str(meeting.stable_key)


def test_explicit_availability_rows_create_non_assumed_profiles() -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    instructor_headers = [cell.value for cell in workbook["Instructors"][1]]
    room_headers = [cell.value for cell in workbook["Rooms"][1]]
    workbook["Instructors"].cell(2, instructor_headers.index("assume_fully_available") + 1, False)
    workbook["Rooms"].cell(2, room_headers.index("assume_fully_available") + 1, False)
    workbook["InstructorAvailability"].append(["FAC-001", "Monday", 0, True])
    workbook["InstructorAvailability"].append(["FAC-001", "Monday", 1, False])
    workbook["RoomAvailability"].append(["CLAB-1", 0, 0, True])
    workbook["RoomAvailability"].append(["CLAB-1", 0, 1, True])
    batch = preview_workbook(workbook_bytes(workbook), term, user)
    assert batch.status == models.ImportStatus.PREVIEWED
    revision = commit_import(batch, user)
    instructor_profile = revision.instructor_availability_profiles.get()
    room_profile = revision.room_availability_profiles.get()
    assert instructor_profile.assume_fully_available is False
    assert instructor_profile.availability_rows.count() == 2
    assert room_profile.assume_fully_available is False
    assert room_profile.availability_rows.count() == 2


def test_preview_reports_formula_header_reference_type_and_completeness_errors() -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    workbook["Instructors"]["B2"] = "=CONCAT(\"Faculty\", \" One\")"
    workbook["Sections"]["C2"] = "not-an-integer"
    workbook["OfferingInstructors"]["B2"] = "UNKNOWN"
    workbook["Subjects"]["A1"] = "wrong_header"
    del workbook["Capabilities"]
    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.INVALID
    codes = set(batch.errors.values_list("code", flat=True))
    assert {
        "FORMULA_NOT_ALLOWED",
        "INVALID_VALUE",
        "UNKNOWN_REFERENCE",
        "MISSING_HEADER",
        "UNEXPECTED_HEADER",
        "MISSING_SHEET",
    }.issubset(codes)
    formula = batch.errors.get(code="FORMULA_NOT_ALLOWED")
    assert formula.sheet_name == "Instructors"
    assert formula.row_number == 2
    assert formula.column_name == "display_name"


def test_preview_rejects_missing_availability_acknowledgement_and_invalid_units() -> None:
    term = make_term()
    user = make_user()
    workbook = valid_workbook(make_policy(term, user))
    workbook["Instructors"]["E2"] = False
    workbook["Rooms"]["H2"] = False
    workbook["Rooms"]["E2"] = "CSM"  # both a college and department owner are now supplied
    workbook["RoomAuthorizations"]["C2"] = "CSM"  # two authorization targets
    batch = preview_workbook(workbook_bytes(workbook), term, user)

    assert batch.status == models.ImportStatus.INVALID
    codes = list(batch.errors.values_list("code", flat=True))
    assert codes.count("MISSING_AVAILABILITY_PROFILE") == 2
    assert "INVALID_OWNER" in codes
    assert "INVALID_AUTHORIZATION_TARGET" in codes


def test_invalid_batch_and_unauthorized_reviewer_cannot_commit() -> None:
    term = make_term()
    central = make_user()
    policy = make_policy(term, central)
    reviewer = models.User.objects.create_user(
        username="reviewer",
        role=models.UserRole.COLLEGE_REVIEWER,
    )
    workbook = valid_workbook(policy)
    del workbook["Sections"]
    invalid = preview_workbook(workbook_bytes(workbook), term, central)
    with pytest.raises(ImportCommitError, match="clean PREVIEWED"):
        commit_import(invalid, central)

    valid = preview_workbook(workbook_bytes(valid_workbook(policy)), term, central)
    with pytest.raises(ImportCommitError, match="Only a system administrator"):
        commit_import(valid, reviewer)
    assert term.dataset_revisions.count() == 0


def test_commit_is_atomic_when_staged_data_is_tampered() -> None:
    term = make_term()
    user = make_user()
    policy = make_policy(term, user)
    batch = preview_workbook(workbook_bytes(valid_workbook(policy)), term, user)
    summary = batch.summary
    summary["sheets"]["Departments"][0]["college_code"] = "DOES-NOT-EXIST"
    batch.summary = summary
    batch.save(update_fields=["summary", "updated_at"])

    with pytest.raises(ImportCommitError, match="failed atomically"):
        commit_import(batch, user)
    assert term.dataset_revisions.count() == 0
    assert models.College.objects.count() == 0
    batch.refresh_from_db()
    assert batch.status == models.ImportStatus.PREVIEWED
    assert batch.committed_revision is None


def test_same_bytes_reuse_preview_and_committed_batch_cannot_commit_twice() -> None:
    term = make_term()
    user = make_user()
    content = workbook_bytes(valid_workbook(make_policy(term, user)))
    first = preview_workbook(content, term, user)
    second = preview_workbook(content, term, user)
    assert second.pk == first.pk

    revision = commit_import(first, user)
    returned = preview_workbook(content, term, user)
    assert returned.pk == first.pk
    assert returned.committed_revision_id == revision.pk
    with pytest.raises(ImportCommitError, match="already been committed"):
        commit_import(first, user)


def test_management_command_writes_template_and_protects_existing_file(tmp_path) -> None:
    destination = tmp_path / "semester-template.xlsx"
    call_command("create_import_template", str(destination))
    assert destination.exists()
    workbook = load_workbook(destination)
    assert workbook[SCHEMA_SHEET]["B1"].value == SCHEMA_VERSION

    with pytest.raises(Exception, match="Refusing to overwrite"):
        call_command("create_import_template", str(destination))
    call_command("create_import_template", str(destination), force=True)
