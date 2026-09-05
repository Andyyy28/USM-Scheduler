from __future__ import annotations

import csv
import hashlib
import json
from io import BytesIO, StringIO

import pytest
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient

from scheduler import models
from scheduler.management.commands.seed_demo import build_demo_workbook_bytes
from scheduler.services.problem_builder import BuildIssue, ProblemBuildError

pytestmark = pytest.mark.django_db


def seed_demo(**options: object) -> dict[str, int]:
    output = StringIO()
    call_command("seed_demo", stdout=output, **options)
    return json.loads(output.getvalue().strip().splitlines()[-1])


def test_inactive_central_user_cannot_retain_privileged_api_access() -> None:
    inactive = models.User.objects.create_user(
        username="disabled-central",
        role=models.UserRole.CENTRAL_SCHEDULER,
        is_active=False,
    )
    client = APIClient()
    client.force_authenticate(inactive)

    assert client.get(reverse("api:import-template")).status_code == 403
    assert client.post(reverse("api:snapshots"), {}, format="json").status_code == 403


def test_snapshot_api_rejects_unapproved_objective_without_persisting() -> None:
    identifiers = seed_demo()
    central = models.User.objects.get(pk=identifiers["central_user_id"])
    revision = models.TermDatasetRevision.objects.get(pk=identifiers["revision_id"])
    objective = models.ObjectiveProfile.objects.create(
        name="Draft objective",
        term=revision.term,
        is_approved=False,
    )
    client = APIClient()
    client.force_authenticate(central)

    response = client.post(
        reverse("api:snapshots"),
        {"revision_id": revision.pk, "objective_profile_id": objective.pk},
        format="json",
    )

    assert response.status_code == 400
    assert response.json()["code"] == "PREFLIGHT_FAILED"
    assert response.json()["detail"]
    assert response.json()["issues"][0]["code"] == "OBJECTIVE_NOT_APPROVED"
    assert not models.ProblemSnapshot.objects.filter(objective_profile=objective).exists()


def test_snapshot_api_preserves_every_structured_preflight_issue(monkeypatch) -> None:  # type: ignore[no-untyped-def]
    identifiers = seed_demo()
    central = models.User.objects.get(pk=identifiers["central_user_id"])
    client = APIClient()
    client.force_authenticate(central)

    def fail_build(*_args, **_kwargs):  # type: ignore[no-untyped-def]
        raise ProblemBuildError(
            [
                BuildIssue("FIRST_ISSUE", "First readable issue."),
                BuildIssue("SECOND_ISSUE", "Second readable issue."),
            ]
        )

    monkeypatch.setattr("scheduler.api.views.build_and_store_snapshot", fail_build)
    response = client.post(
        reverse("api:snapshots"),
        {
            "revision_id": identifiers["revision_id"],
            "objective_profile_id": identifiers["objective_profile_id"],
        },
        format="json",
    )

    assert response.status_code == 400
    assert response.json() == {
        "code": "PREFLIGHT_FAILED",
        "detail": "First readable issue.; Second readable issue.",
        "issues": [
            {"code": "FIRST_ISSUE", "message": "First readable issue.", "entity_type": "", "entity_id": ""},
            {"code": "SECOND_ISSUE", "message": "Second readable issue.", "entity_type": "", "entity_id": ""},
        ],
    }


def test_seed_demo_is_deterministic_idempotent_deidentified_and_password_safe() -> None:
    first_bytes = build_demo_workbook_bytes()
    second_bytes = build_demo_workbook_bytes()
    assert first_bytes == second_bytes
    assert hashlib.sha256(first_bytes).hexdigest() == hashlib.sha256(second_bytes).hexdigest()

    first = seed_demo()
    second = seed_demo()
    assert second == first
    assert models.AcademicTerm.objects.count() == 1
    assert models.TermDatasetRevision.objects.count() == 1
    assert models.ImportBatch.objects.count() == 1
    assert models.ObjectiveProfile.objects.count() == 1

    admin = models.User.objects.get(pk=first["admin_user_id"])
    central = models.User.objects.get(pk=first["central_user_id"])
    reviewer = models.User.objects.get(pk=first["reviewer_user_id"])
    assert admin.has_usable_password() is False
    assert central.has_usable_password() is False
    assert reviewer.has_usable_password() is False
    assert admin.is_superuser and admin.is_staff
    assert reviewer.college_scopes.filter(college_id=first["review_college_id"]).exists()

    pseudonymous_codes = list(models.Student.objects.values_list("pseudonymous_code", flat=True))
    assert pseudonymous_codes == ["demo-anon-001"]
    assert all("student" not in code.casefold() for code in pseudonymous_codes)

    third = seed_demo(central_password="configured-test-secret")
    assert third == first
    central.refresh_from_db()
    assert central.check_password("configured-test-secret")


@override_settings(
    CELERY_TASK_ALWAYS_EAGER=True,
    CELERY_TASK_EAGER_PROPAGATES=True,
    SOLVER_DEFAULT_TIME_LIMIT_SECONDS=1,
)
def test_authenticated_api_full_schedule_workflow() -> None:
    identifiers = seed_demo()
    central = models.User.objects.get(pk=identifiers["central_user_id"])
    reviewer = models.User.objects.get(pk=identifiers["reviewer_user_id"])
    revision = models.TermDatasetRevision.objects.get(pk=identifiers["revision_id"])
    objective = models.ObjectiveProfile.objects.get(pk=identifiers["objective_profile_id"])
    college = models.College.objects.get(pk=identifiers["review_college_id"])
    client = APIClient()

    assert client.get(reverse("api:terms")).status_code in {401, 403}
    client.force_authenticate(central)
    template_response = client.get(reverse("api:import-template"))
    assert template_response.status_code == 200
    assert template_response.content.startswith(b"PK")
    assert "spreadsheetml" in template_response["Content-Type"]

    term_response = client.get(reverse("api:terms"))
    assert term_response.status_code == 200
    assert term_response.json()[0]["id"] == identifiers["term_id"]
    revision_response = client.get(
        reverse("api:revisions", kwargs={"term_id": identifiers["term_id"]})
    )
    assert revision_response.status_code == 200
    revision_payload = revision_response.json()[0]
    assert revision_payload["status"] == models.RevisionStatus.COMMITTED
    assert revision_payload["data_origin"] == models.DatasetOrigin.SYNTHETIC
    assert revision_payload["data_origin_display"] == "Synthetic / practice"
    assert revision_payload["source_filename"]
    assert revision_payload["section_count"] == 1
    assert revision_payload["meeting_count"] == 2
    assert revision_payload["room_count"] == 2
    assert revision_payload["instructor_count"] == 2

    snapshot_response = client.post(
        reverse("api:snapshots"),
        {
            "revision_id": revision.pk,
            "objective_profile_id": objective.pk,
        },
        format="json",
    )
    assert snapshot_response.status_code == 201, snapshot_response.content
    snapshot_payload = snapshot_response.json()
    snapshot = models.ProblemSnapshot.objects.get(pk=snapshot_payload["id"])
    assert snapshot_payload["event_count"] == 2
    assert snapshot_payload["candidate_count"] > 0
    assert snapshot_payload["snapshot_hash"] == snapshot.snapshot_hash
    run_response = client.post(
        reverse("api:runs"),
        {
            "snapshot_id": snapshot.pk,
            "algorithms": [models.SolverAlgorithm.CP_SAT, models.SolverAlgorithm.GENETIC_ALGORITHM],
            "seed": 1001,
            "configuration": {
                "time_limit_seconds": 1,
                "worker_count": 1,
                "population_size": 24,
                "tournament_size": 3,
                "elite_fraction": 0.1,
                "repair_attempts": 8,
                "max_generations": 30,
            },
        },
        format="json",
    )
    assert run_response.status_code == 202, run_response.content
    run_payloads = run_response.json()
    assert {row["algorithm"] for row in run_payloads} == set(models.SolverAlgorithm.values)
    runs = list(models.ScheduleRun.objects.filter(snapshot=snapshot).order_by("algorithm"))
    assert len(runs) == 2
    assert all(run.status in {models.RunStatus.FEASIBLE, models.RunStatus.OPTIMAL} for run in runs)
    assert all(run.snapshot.snapshot_hash == snapshot.snapshot_hash for run in runs)
    assert len({run.diagnostics["problem_hash"] for run in runs}) == 1
    assert all(run.validation_result.is_feasible for run in runs)
    assert all(run.schedule_version.assignments.count() == 2 for run in runs)

    comparison = client.get(reverse("api:run-comparison"), {"snapshot_id": snapshot.pk})
    assert comparison.status_code == 200
    assert comparison.json()[models.SolverAlgorithm.CP_SAT]["feasible_runs"] == 1
    assert comparison.json()[models.SolverAlgorithm.GENETIC_ALGORITHM]["feasible_runs"] == 1

    schedule = models.ScheduleVersion.objects.get(run__algorithm=models.SolverAlgorithm.CP_SAT)
    draft_export = client.get(
        reverse(
            "api:schedule-export",
            kwargs={"schedule_id": schedule.pk, "export_format": "csv"},
        )
    )
    assert draft_export.status_code == 400
    validation_response = client.post(
        reverse("api:schedule-validate", kwargs={"schedule_id": schedule.pk}),
        {},
        format="json",
    )
    assert validation_response.status_code == 200
    assert validation_response.json()["feasible"] is True

    submit_response = client.post(
        reverse("api:schedule-submit-review", kwargs={"schedule_id": schedule.pk}),
        {},
        format="json",
    )
    assert submit_response.status_code == 200, submit_response.content
    schedule.refresh_from_db()
    assert schedule.status == models.ScheduleStatus.UNDER_REVIEW

    client.force_authenticate(reviewer)
    assert client.get(reverse("api:schedule-detail", kwargs={"schedule_id": schedule.pk})).status_code == 200
    assert client.get(reverse("api:import-template")).status_code == 403
    assert (
        client.post(
            reverse("api:snapshots"),
            {"revision_id": revision.pk, "objective_profile_id": objective.pk},
            format="json",
        ).status_code
        == 403
    )
    assert (
        client.post(
            reverse("api:schedule-approve", kwargs={"schedule_id": schedule.pk}),
            {},
            format="json",
        ).status_code
        == 403
    )

    outside_college = models.College.objects.create(code="OUT", name="Out-of-scope College")
    out_of_scope = client.post(
        reverse("api:schedule-review", kwargs={"schedule_id": schedule.pk}),
        {
            "college_id": outside_college.pk,
            "status": models.ReviewStatus.ENDORSED,
            "comment": "Should be rejected",
        },
        format="json",
    )
    assert out_of_scope.status_code == 403

    endorsement = client.post(
        reverse("api:schedule-review", kwargs={"schedule_id": schedule.pk}),
        {
            "college_id": college.pk,
            "status": models.ReviewStatus.ENDORSED,
            "comment": "Reviewed against the college room policy.",
        },
        format="json",
    )
    assert endorsement.status_code == 201, endorsement.content
    assert endorsement.json()["status"] == models.ReviewStatus.ENDORSED

    client.force_authenticate(central)
    approval_response = client.post(
        reverse("api:schedule-approve", kwargs={"schedule_id": schedule.pk}),
        {"notes": "Approved demonstration schedule"},
        format="json",
    )
    assert approval_response.status_code == 200, approval_response.content
    schedule.refresh_from_db()
    assert schedule.status == models.ScheduleStatus.APPROVED
    assert schedule.approval.approved_by == central

    assignment_ids = list(schedule.assignments.values_list("pk", flat=True))
    lock_response = client.post(
        reverse("api:schedule-lock", kwargs={"schedule_id": schedule.pk}),
        {
            "assignment_ids": assignment_ids,
            "reason": "Carry approved placements into regeneration",
        },
        format="json",
    )
    assert lock_response.status_code == 201, lock_response.content
    assert len(lock_response.json()["lock_ids"]) == len(assignment_ids)
    assert models.LockedAssignment.objects.filter(source_schedule=schedule, is_active=True).count() == 2

    csv_response = client.get(
        reverse(
            "api:schedule-export",
            kwargs={"schedule_id": schedule.pk, "export_format": "csv"},
        )
    )
    assert csv_response.status_code == 200
    assert csv_response["Content-Type"].startswith("text/csv")
    assert csv_response["Content-Disposition"].endswith('.csv"')
    csv_text = csv_response.content.decode("utf-8-sig")
    csv_rows = list(csv.DictReader(StringIO(csv_text)))
    assert len(csv_rows) == 2
    assert {
        "meeting_id",
        "offering_key",
        "subject_code",
        "sections",
        "instructors",
        "day",
        "starts_at",
        "ends_at",
        "room_code",
        "offering_unit",
        "locked",
    }.issubset(csv_rows[0])
    assert {row["locked"] for row in csv_rows} == {"YES"}
    assert "demo-anon-001" not in csv_text

    xlsx_response = client.get(
        reverse(
            "api:schedule-export",
            kwargs={"schedule_id": schedule.pk, "export_format": "xlsx"},
        )
    )
    assert xlsx_response.status_code == 200
    assert xlsx_response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert xlsx_response["Content-Disposition"].endswith('.xlsx"')
    assert xlsx_response.content.startswith(b"PK")
    exported_workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True)
    assert {"Schedule", "Manifest", "Validation"}.issubset(exported_workbook.sheetnames)
    schedule_values = [
        str(cell.value)
        for row in exported_workbook["Schedule"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    all_exported_values = [
        str(cell.value)
        for sheet in exported_workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "Meeting Id" in schedule_values
    assert "demo-anon-001" not in "\n".join(all_exported_values)

    schedule.name = "Forbidden rewrite"
    with pytest.raises(ValidationError, match="immutable"):
        schedule.save()

    second_schedule = models.ScheduleVersion.objects.get(
        run__algorithm=models.SolverAlgorithm.GENETIC_ALGORITHM
    )
    assert second_schedule.status == models.ScheduleStatus.DRAFT
    assert not hasattr(second_schedule, "approval")
